from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, date, time
from functools import wraps
import os
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import seaborn as sns
import base64
import json
from dotenv import load_dotenv
import csv
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Load environment variables
load_dotenv()

app = Flask(__name__)

# Import configuration after creating the app
try:
    from config import config
    config_name = os.environ.get('FLASK_ENV', 'development')
    app.config.from_object(config[config_name])
except ImportError:
    # Fallback configuration if config.py doesn't exist
    app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-here')
    app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///instance/crm.db')
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Add nl2br filter for templates
@app.template_filter('nl2br')
def nl2br(value):
    """Convert newlines to <br> tags"""
    if value:
        return value.replace('\n', '<br>\n')
    return value

# Helper functions to replace pandas functionality
def is_empty_value(value):
    """Check if a value is empty, None, or NaN equivalent"""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == '':
        return True
    if isinstance(value, (int, float)) and str(value).lower() in ['nan', 'null', '']:
        return True
    return False

def safe_str(value, default=''):
    """Safely convert value to string, handling None and empty values"""
    if is_empty_value(value):
        return default
    return str(value).strip()

def create_excel_from_data(data, filename, sheet_name='Sheet1'):
    """Create Excel file from data using openpyxl"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    if data:
        # Add headers
        headers = list(data[0].keys())
        ws.append(headers)
        
        # Add data rows
        for row_data in data:
            row = [row_data.get(col, '') for col in headers]
            ws.append(row)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def read_excel_file(file_path):
    """Read Excel file and return data as list of dictionaries"""
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value.strip() if cell.value else '')
    
    # Get data rows
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = value
        data.append(row_dict)
    
    return data

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Database Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # admin, instructor, customer_service
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)

# Association table for many-to-many relationship between customers and instructors
customer_instructor_association = db.Table('customer_instructor',
    db.Column('customer_id', db.Integer, db.ForeignKey('customer.id'), primary_key=True),
    db.Column('instructor_id', db.Integer, db.ForeignKey('user.id'), primary_key=True),
    db.Column('assigned_at', db.DateTime, default=datetime.utcnow)
)

class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    phone = db.Column(db.String(20))
    phone2 = db.Column(db.String(20))  # Second phone number
    age = db.Column(db.Integer)  # Customer age
    status = db.Column(db.String(20), default='active')  # active, inactive, needs_follow_up, no_show
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'))  # Track who added the customer
    initial_notes = db.Column(db.Text)
    
    # Many-to-many relationship with instructors
    assigned_instructors = db.relationship('User', 
                                         secondary=customer_instructor_association,
                                         backref=db.backref('assigned_customers', lazy='dynamic'),
                                         lazy='dynamic')
    
    # Relationship to track who created the customer
    created_by = db.relationship('User', foreign_keys=[created_by_id], backref='customers_created')

class Course(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    instructor_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    status = db.Column(db.String(20), default='open')  # open, in_progress, completed, cancelled
    scheduled_date = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    completion_notes = db.Column(db.Text)
    
    customer = db.relationship('Customer', backref='courses')
    instructor = db.relationship('User', backref='assigned_courses')

class Ticket(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text, nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    assigned_to_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    status = db.Column(db.String(20), default='open')  # open, in_progress, resolved, closed
    priority = db.Column(db.String(10), default='medium')  # low, medium, high, urgent
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    resolved_at = db.Column(db.DateTime)
    resolution_notes = db.Column(db.Text)
    
    customer = db.relationship('Customer', backref='tickets')
    assigned_to = db.relationship('User', backref='assigned_tickets')

class Session(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    instructor_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    course_id = db.Column(db.Integer, db.ForeignKey('course.id'))
    scheduled_date = db.Column(db.DateTime, nullable=False)
    duration = db.Column(db.Integer, default=60)  # minutes
    status = db.Column(db.String(20), default='scheduled')  # scheduled, completed, no_show, cancelled
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    customer = db.relationship('Customer', backref='sessions')
    instructor = db.relationship('User', backref='instructor_sessions')
    course = db.relationship('Course', backref='sessions')

class Note(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    content = db.Column(db.Text, nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    is_internal = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    customer = db.relationship('Customer', backref='notes_list')
    created_by = db.relationship('User', backref='created_notes')

# New Course Management Models
class CourseCategory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    description = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)



# Group Management Models
class Group(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    subject = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    instructor_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    category_id = db.Column(db.Integer, db.ForeignKey('course_category.id'))
    max_students = db.Column(db.Integer, default=15)
    status = db.Column(db.String(20), default='active')  # active, inactive, completed
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    notes = db.Column(db.Text)
    
    instructor = db.relationship('User', backref='instructor_groups')
    category = db.relationship('CourseCategory', backref='category_groups')

class GroupSchedule(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey('group.id'), nullable=False)
    day_of_week = db.Column(db.String(10), nullable=False)  # monday, tuesday, etc.
    start_time = db.Column(db.Time, nullable=False)
    end_time = db.Column(db.Time, nullable=False)
    location = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    group = db.relationship('Group', backref='schedules')

class GroupMember(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey('group.id'), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    joined_date = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default='active')  # active, inactive, completed, dropped
    notes = db.Column(db.Text)
    
    group = db.relationship('Group', backref='members')
    customer = db.relationship('Customer', backref='group_memberships')

class GroupSession(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey('group.id'), nullable=False)
    session_date = db.Column(db.Date, nullable=False)
    start_time = db.Column(db.Time, nullable=False)
    end_time = db.Column(db.Time, nullable=False)
    status = db.Column(db.String(20), default='scheduled')  # scheduled, completed, cancelled
    topic = db.Column(db.String(200))
    notes = db.Column(db.Text)
    location = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    group = db.relationship('Group', backref='group_sessions')

class GroupAttendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    group_session_id = db.Column(db.Integer, db.ForeignKey('group_session.id'), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    status = db.Column(db.String(20), default='present')  # present, absent, late, excused
    notes = db.Column(db.Text)
    recorded_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    group_session = db.relationship('GroupSession', backref='attendance_records')
    customer = db.relationship('Customer', backref='attendance_records')

# New audit trail and history models
class AuditLog(db.Model):
    """Track all important actions in the system"""
    id = db.Column(db.Integer, primary_key=True)
    entity_type = db.Column(db.String(50), nullable=False)  # 'group', 'customer', 'session', etc.
    entity_id = db.Column(db.Integer, nullable=False)
    action = db.Column(db.String(50), nullable=False)  # 'created', 'updated', 'deleted', 'member_added', etc.
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    old_values = db.Column(db.Text)  # JSON string of old values
    new_values = db.Column(db.Text)  # JSON string of new values
    description = db.Column(db.String(255))
    ip_address = db.Column(db.String(45))
    user_agent = db.Column(db.String(255))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    user = db.relationship('User', backref='audit_logs')

class CustomerHistory(db.Model):
    """Track customer interactions and changes"""
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    event_type = db.Column(db.String(50), nullable=False)  # 'status_change', 'group_joined', 'session_attended', etc.
    event_description = db.Column(db.String(255), nullable=False)
    event_data = db.Column(db.Text)  # JSON string for additional data
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    customer = db.relationship('Customer', backref='history_events')
    created_by = db.relationship('User', backref='customer_events_created')

class GroupHistory(db.Model):
    """Track group events and changes"""
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey('group.id'), nullable=False)
    event_type = db.Column(db.String(50), nullable=False)  # 'created', 'member_added', 'session_completed', etc.
    event_description = db.Column(db.String(255), nullable=False)
    event_data = db.Column(db.Text)  # JSON string for additional data
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    group = db.relationship('Group', backref='history_events')
    created_by = db.relationship('User', backref='group_events_created')

class Performance(db.Model):
    """Track student performance metrics"""
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    group_id = db.Column(db.Integer, db.ForeignKey('group.id'), nullable=False)
    assessment_date = db.Column(db.Date, nullable=False)
    score = db.Column(db.Float)  # 0-100 scale
    grade = db.Column(db.String(10))  # A, B, C, D, F or custom grading
    assessment_type = db.Column(db.String(50))  # 'quiz', 'exam', 'project', 'participation'
    notes = db.Column(db.Text)
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    customer = db.relationship('Customer', backref='performance_records')
    group = db.relationship('Group', backref='performance_records')
    created_by = db.relationship('User', backref='assessments_created')

class Communication(db.Model):
    """Track all communications with customers"""
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    communication_type = db.Column(db.String(50), nullable=False)  # 'phone', 'whatsapp', 'email', 'in_person'
    direction = db.Column(db.String(20), nullable=False)  # 'outbound', 'inbound'
    subject = db.Column(db.String(200))
    content = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(20), default='completed')  # 'completed', 'failed', 'pending'
    response_received = db.Column(db.Boolean, default=False)
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    customer = db.relationship('Customer', backref='communications')
    created_by = db.relationship('User', backref='communications_sent')

# Helper function to log activities
def log_activity(entity_type, entity_id, action, user_id, description=None, old_values=None, new_values=None):
    """Helper function to create audit log entries"""
    audit_log = AuditLog(
        entity_type=entity_type,
        entity_id=entity_id,
        action=action,
        user_id=user_id,
        description=description,
        old_values=json.dumps(old_values) if old_values else None,
        new_values=json.dumps(new_values) if new_values else None,
        ip_address=request.remote_addr if request else None,
        user_agent=request.headers.get('User-Agent') if request else None
    )
    db.session.add(audit_log)

def log_customer_event(customer_id, event_type, description, event_data=None, created_by_id=None):
    """Helper function to log customer events"""
    event = CustomerHistory(
        customer_id=customer_id,
        event_type=event_type,
        event_description=description,
        event_data=json.dumps(event_data) if event_data else None,
        created_by_id=created_by_id or current_user.id
    )
    db.session.add(event)

def log_group_event(group_id, event_type, description, event_data=None, created_by_id=None):
    """Helper function to log group events"""
    event = GroupHistory(
        group_id=group_id,
        event_type=event_type,
        event_description=description,
        event_data=json.dumps(event_data) if event_data else None,
        created_by_id=created_by_id or current_user.id
    )
    db.session.add(event)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def role_required(role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated or current_user.role != role:
                flash('Access denied. Insufficient permissions.', 'error')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Admin access required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# Routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password) and user.is_active:
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.role == 'admin':
        return render_template('admin_dashboard.html')
    elif current_user.role == 'instructor':
        return render_template('instructor_dashboard.html')
    elif current_user.role == 'customer_service':
        return render_template('customer_service_dashboard.html')
    else:
        flash('Invalid user role', 'error')
        return redirect(url_for('logout'))

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        # Update user profile
        current_user.first_name = request.form['first_name']
        current_user.last_name = request.form['last_name']
        current_user.email = request.form['email']
        
        # Update password if provided
        if request.form.get('current_password') and request.form.get('new_password'):
            if check_password_hash(current_user.password_hash, request.form['current_password']):
                if request.form['new_password'] == request.form['confirm_password']:
                    current_user.password_hash = generate_password_hash(request.form['new_password'])
                    flash('Profile and password updated successfully!', 'success')
                else:
                    flash('New passwords do not match!', 'error')
                    return render_template('profile.html')
            else:
                flash('Current password is incorrect!', 'error')
                return render_template('profile.html')
        else:
            flash('Profile updated successfully!', 'success')
        
        db.session.commit()
        return redirect(url_for('profile'))
    
    return render_template('profile.html')

# Admin Routes
@app.route('/admin/users')
@admin_required
def admin_users():
    users = User.query.all()
    return render_template('admin/users.html', users=users)

@app.route('/admin/users/add', methods=['GET', 'POST'])
@admin_required
def admin_add_user():
    if request.method == 'POST':
        user = User(
            username=request.form['username'],
            email=request.form['email'],
            password_hash=generate_password_hash(request.form['password']),
            role=request.form['role'],
            first_name=request.form['first_name'],
            last_name=request.form['last_name']
        )
        db.session.add(user)
        
        try:
            db.session.commit()
            flash('User added successfully', 'success')
            return redirect(url_for('admin_users'))
        except Exception as e:
            db.session.rollback()
            flash('Error adding user: Username or email may already exist', 'error')
    
    return render_template('admin/add_user.html')

@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_edit_user(user_id):
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        user.username = request.form['username']
        user.email = request.form['email']
        user.first_name = request.form['first_name']
        user.last_name = request.form['last_name']
        user.role = request.form['role']
        user.is_active = 'is_active' in request.form
        
        # Update password if provided
        if request.form.get('password'):
            user.password_hash = generate_password_hash(request.form['password'])
        
        try:
            db.session.commit()
            flash('User updated successfully', 'success')
            return redirect(url_for('admin_users'))
        except Exception as e:
            db.session.rollback()
            flash('Error updating user: Username or email may already exist', 'error')
    
    return render_template('admin/edit_user.html', user=user)

@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def admin_delete_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # Prevent deleting the current admin user
    if user.id == current_user.id:
        flash('You cannot delete your own account', 'error')
        return redirect(url_for('admin_users'))
    
    # Check if user has assigned customers or tickets
    if user.assigned_customers or user.assigned_tickets or user.instructor_sessions:
        flash('Cannot delete user with assigned customers, tickets, or sessions. Please reassign them first.', 'error')
        return redirect(url_for('admin_users'))
    
    try:
        db.session.delete(user)
        db.session.commit()
        flash(f'User {user.username} has been deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error deleting user', 'error')
    
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/toggle', methods=['POST'])
@admin_required
def admin_toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # Prevent toggling the current admin user
    if user.id == current_user.id:
        flash('You cannot deactivate your own account', 'error')
        return redirect(url_for('admin_users'))
    
    user.is_active = not user.is_active
    status = 'activated' if user.is_active else 'deactivated'
    
    try:
        db.session.commit()
        flash(f'User {user.username} has been {status}', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error updating user status', 'error')
    
    return redirect(url_for('admin_users'))

@app.route('/admin/reports')
@admin_required
def admin_reports():
    # Calculate statistics
    total_customers = Customer.query.count()
    total_groups = Group.query.count()
    total_tickets = Ticket.query.count()
    open_tickets = Ticket.query.filter_by(status='open').count()
    resolved_tickets = Ticket.query.filter_by(status='resolved').count()
    
    # Recent activities
    recent_customers = Customer.query.order_by(Customer.created_at.desc()).limit(5).all()
    recent_tickets = Ticket.query.order_by(Ticket.created_at.desc()).limit(5).all()
    
    stats = {
        'total_customers': total_customers,
        'total_groups': total_groups,
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'resolved_tickets': resolved_tickets
    }
    
    return render_template('admin/reports.html', stats=stats, 
                         recent_customers=recent_customers, recent_tickets=recent_tickets)

# Customer Management Routes
@app.route('/customers')
@login_required
def customers():
    # Instructors can only see their students (customers in their groups)
    if current_user.role == 'instructor':
        # Get customers who are members of instructor's groups
        instructor_groups = Group.query.filter_by(instructor_id=current_user.id).all()
        group_ids = [group.id for group in instructor_groups]
        
        if group_ids:
            # Get customers who are active members of these groups
            customer_ids = db.session.query(GroupMember.customer_id).filter(
                GroupMember.group_id.in_(group_ids),
                GroupMember.status == 'active'
            ).distinct().all()
            customer_ids = [cid[0] for cid in customer_ids]
            
            if customer_ids:
                customers = Customer.query.filter(Customer.id.in_(customer_ids)).all()
            else:
                customers = []
        else:
            customers = []
        
        # For instructors, we don't need complex filtering
        return render_template('customers.html', 
                             customers=customers,
                             customer_service_staff=[],  # Empty for instructors
                             age_filter=None, 
                             min_age=None, 
                             max_age=None,
                             created_by_filter=None,
                             enrollment_status_filter=None)
    
    # Get base query
    query = Customer.query
    
    # Apply filters
    age_filter = request.args.get('age_filter')
    min_age = request.args.get('min_age')
    max_age = request.args.get('max_age')
    created_by_filter = request.args.get('created_by_filter')
    enrollment_status_filter = request.args.get('enrollment_status_filter')
    
    # Age filtering
    if age_filter:
        if age_filter == 'under_18':
            query = query.filter(Customer.age < 18)
        elif age_filter == '18_25':
            query = query.filter(Customer.age >= 18, Customer.age <= 25)
        elif age_filter == '26_35':
            query = query.filter(Customer.age >= 26, Customer.age <= 35)
        elif age_filter == '36_50':
            query = query.filter(Customer.age >= 36, Customer.age <= 50)
        elif age_filter == 'over_50':
            query = query.filter(Customer.age > 50)
        elif age_filter == 'unknown':
            query = query.filter(Customer.age.is_(None))
    elif min_age or max_age:
        # Custom age range
        if min_age and min_age.isdigit():
            query = query.filter(Customer.age >= int(min_age))
        if max_age and max_age.isdigit():
            query = query.filter(Customer.age <= int(max_age))
    
    # Filter by who created the customer
    if created_by_filter and created_by_filter.isdigit():
        query = query.filter(Customer.created_by_id == int(created_by_filter))
    
    # Filter by enrollment status
    if enrollment_status_filter:
        if enrollment_status_filter == 'enrolled':
            # Customers who are enrolled in groups
            query = query.filter(Customer.group_memberships.any(GroupMember.status == 'active'))
        elif enrollment_status_filter == 'not_enrolled':
            # Customers who are not enrolled in any active groups
            query = query.filter(~Customer.group_memberships.any(GroupMember.status == 'active'))
        elif enrollment_status_filter == 'has_courses':
            # Customers who have individual courses
            query = query.filter(Customer.courses.any())
        elif enrollment_status_filter == 'no_courses':
            # Customers who have no individual courses
            query = query.filter(~Customer.courses.any())
    
    customers = query.all()
    
    # Get all customer service staff for the filter dropdown
    customer_service_staff = User.query.filter(
        User.role.in_(['admin', 'customer_service'])
    ).order_by(User.first_name, User.last_name).all()
    
    return render_template('customers.html', 
                         customers=customers,
                         customer_service_staff=customer_service_staff,
                         age_filter=age_filter, 
                         min_age=min_age, 
                         max_age=max_age,
                         created_by_filter=created_by_filter,
                         enrollment_status_filter=enrollment_status_filter)

@app.route('/customers/add', methods=['GET', 'POST'])
@login_required
def add_customer():
    # منع المدرسين من إضافة عملاء
    if current_user.role == 'instructor':
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        # Handle age conversion
        age = None
        age_input = request.form.get('age', '').strip()
        if age_input and age_input.isdigit():
            age = int(age_input)
        
        customer = Customer(
            first_name=request.form['first_name'],
            last_name=request.form['last_name'],
            phone=request.form.get('phone', ''),
            phone2=request.form.get('phone2', ''),
            age=age,
            initial_notes=request.form.get('notes', ''),
            created_by_id=current_user.id
        )
        db.session.add(customer)
        db.session.flush()  # Get the customer ID
        
        # Handle multiple instructor assignments
        instructor_ids = request.form.getlist('assigned_instructor_ids')
        if instructor_ids:
            for instructor_id in instructor_ids:
                if instructor_id:  # Skip empty values
                    instructor = User.query.get(instructor_id)
                    if instructor and instructor.role == 'instructor':
                        customer.assigned_instructors.append(instructor)
        
        # Handle group assignments
        group_ids = request.form.getlist('group_ids')
        if group_ids:
            for group_id in group_ids:
                if group_id:  # Skip empty values
                    group = Group.query.get(group_id)
                    if group and group.status == 'active':
                        # Check if group is not full
                        current_members = GroupMember.query.filter_by(
                            group_id=group.id, 
                            status='active'
                        ).count()
                        
                        if current_members < group.max_students:
                            # Add customer to group
                            group_member = GroupMember(
                                group_id=group.id,
                                customer_id=customer.id,
                                status='active'
                            )
                            db.session.add(group_member)
                            
                            # Log this activity
                            log_customer_event(
                                customer_id=customer.id,
                                event_type='group_joined',
                                description=f'Enrolled in group: {group.name}',
                                event_data={'group_id': group.id, 'group_name': group.name},
                                created_by_id=current_user.id
                            )
                            
                            log_group_event(
                                group_id=group.id,
                                event_type='member_added',
                                description=f'Student {customer.first_name} {customer.last_name} enrolled',
                                event_data={'customer_id': customer.id},
                                created_by_id=current_user.id
                            )
        
        db.session.commit()
        flash('Customer added successfully and enrolled in selected groups', 'success')
        return redirect(url_for('customers'))
    
    instructors = User.query.filter_by(role='instructor').all()
    # Get active groups for enrollment
    groups = Group.query.filter_by(status='active').order_by(Group.name).all()
    return render_template('add_customer.html', instructors=instructors, groups=groups)

@app.route('/customers/<int:customer_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_customer(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    
    # Check permissions
    if current_user.role == 'instructor' and current_user not in customer.assigned_instructors:
        flash('Access denied', 'error')
        return redirect(url_for('customers'))
    
    if request.method == 'POST':
        customer.first_name = request.form['first_name']
        customer.last_name = request.form['last_name']
        customer.phone = request.form.get('phone', '')
        customer.phone2 = request.form.get('phone2', '')
        
        # Handle age conversion
        age_input = request.form.get('age', '').strip()
        if age_input and age_input.isdigit():
            customer.age = int(age_input)
        else:
            customer.age = None
            
        customer.status = request.form['status']
        
        # Only admins and customer service can change instructor assignment
        if current_user.role in ['admin', 'customer_service']:
            # Clear existing assignments
            customer.assigned_instructors = []
            
            # Add new assignments
            instructor_ids = request.form.getlist('assigned_instructor_ids')
            if instructor_ids:
                for instructor_id in instructor_ids:
                    if instructor_id:  # Skip empty values
                        instructor = User.query.get(instructor_id)
                        if instructor and instructor.role == 'instructor':
                            customer.assigned_instructors.append(instructor)
        
        # Handle group management
        if current_user.role in ['admin', 'customer_service', 'instructor']:
            # Remove from selected groups
            remove_group_ids = request.form.getlist('remove_group_ids')
            for group_id in remove_group_ids:
                if group_id:
                    membership = GroupMember.query.filter_by(
                        customer_id=customer.id,
                        group_id=int(group_id),
                        status='active'
                    ).first()
                    if membership:
                        membership.status = 'dropped'
                        group = Group.query.get(group_id)
                        
                        # Log this activity
                        log_customer_event(
                            customer_id=customer.id,
                            event_type='group_left',
                            description=f'Removed from group: {group.name}',
                            event_data={'group_id': group.id, 'group_name': group.name},
                            created_by_id=current_user.id
                        )
                        
                        log_group_event(
                            group_id=group.id,
                            event_type='member_removed',
                            description=f'Student {customer.first_name} {customer.last_name} removed',
                            event_data={'customer_id': customer.id},
                            created_by_id=current_user.id
                        )
            
            # Add to new groups
            add_group_ids = request.form.getlist('add_group_ids')
            for group_id in add_group_ids:
                if group_id:
                    group = Group.query.get(group_id)
                    if group and group.status == 'active':
                        # Check if not already a member
                        existing_membership = GroupMember.query.filter_by(
                            customer_id=customer.id,
                            group_id=group.id,
                            status='active'
                        ).first()
                        
                        if not existing_membership:
                            # Check if group is not full
                            current_members = GroupMember.query.filter_by(
                                group_id=group.id,
                                status='active'
                            ).count()
                            
                            if current_members < group.max_students:
                                # Add customer to group
                                group_member = GroupMember(
                                    group_id=group.id,
                                    customer_id=customer.id,
                                    status='active'
                                )
                                db.session.add(group_member)
                                
                                # Log this activity
                                log_customer_event(
                                    customer_id=customer.id,
                                    event_type='group_joined',
                                    description=f'Enrolled in group: {group.name}',
                                    event_data={'group_id': group.id, 'group_name': group.name},
                                    created_by_id=current_user.id
                                )
                                
                                log_group_event(
                                    group_id=group.id,
                                    event_type='member_added',
                                    description=f'Student {customer.first_name} {customer.last_name} enrolled',
                                    event_data={'customer_id': customer.id},
                                    created_by_id=current_user.id
                                )
        
        try:
            db.session.commit()
            flash('Customer updated successfully', 'success')
            return redirect(url_for('customers'))
        except Exception as e:
            db.session.rollback()
            flash('Error updating customer', 'error')
    
    instructors = User.query.filter_by(role='instructor').all()
    
    # Get available groups (not already enrolled in)
    customer_group_ids = [m.group_id for m in customer.group_memberships if m.status == 'active']
    available_groups_query = Group.query.filter_by(status='active').filter(
        ~Group.id.in_(customer_group_ids)
    )
    
    available_groups = []
    for group in available_groups_query.all():
        current_members = GroupMember.query.filter_by(
            group_id=group.id,
            status='active'
        ).count()
        group.current_members = current_members
        available_groups.append(group)
    
    return render_template('edit_customer.html', 
                         customer=customer, 
                         instructors=instructors,
                         available_groups=available_groups)

@app.route('/customers/<int:customer_id>/add_note', methods=['POST'])
@login_required
def add_customer_note(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    
    # Check permissions
    if current_user.role == 'instructor' and current_user not in customer.assigned_instructors:
        flash('Access denied', 'error')
        return redirect(url_for('customers'))
    
    content = request.form.get('note_content')
    is_internal = 'is_internal' in request.form
    
    if content:
        note = Note(
            content=content,
            customer_id=customer_id,
            created_by_id=current_user.id,
            is_internal=is_internal
        )
        db.session.add(note)
        db.session.commit()
        flash('Note added successfully', 'success')
    else:
        flash('Note content cannot be empty', 'error')
    
    return redirect(url_for('customer_detail', customer_id=customer_id))

@app.route('/customers/<int:customer_id>/delete', methods=['POST'])
@login_required
def delete_customer(customer_id):
    # Only allow admin and customer_service roles to delete customers
    if current_user.role not in ['admin', 'customer_service']:
        flash('Access denied', 'error')
        return redirect(url_for('customers'))
    
    customer = Customer.query.get_or_404(customer_id)
    
    # Check if customer has any related data that would prevent deletion
    has_tickets = bool(customer.tickets)
    has_sessions = bool(customer.sessions)
    has_courses = bool(customer.courses)
    has_group_memberships = bool(customer.group_memberships)
    has_notes = bool(customer.notes_list)
    has_communications = bool(customer.communications)
    has_performance_records = bool(customer.performance_records)
    has_attendance_records = bool(customer.attendance_records)
    has_history_events = bool(customer.history_events)
    
    # If customer has any related data, prevent deletion
    if (has_tickets or has_sessions or has_courses or has_group_memberships or 
        has_notes or has_communications or has_performance_records or 
        has_attendance_records or has_history_events):
        
        related_items = []
        if has_tickets:
            related_items.append(f"{len(customer.tickets)} ticket(s)")
        if has_sessions:
            related_items.append(f"{len(customer.sessions)} session(s)")
        if has_courses:
            related_items.append(f"{len(customer.courses)} course(s)")
        if has_group_memberships:
            related_items.append(f"{len(customer.group_memberships)} group membership(s)")
        if has_notes:
            related_items.append(f"{len(customer.notes_list)} note(s)")
        if has_communications:
            related_items.append(f"{len(customer.communications)} communication(s)")
        if has_performance_records:
            related_items.append(f"{len(customer.performance_records)} performance record(s)")
        if has_attendance_records:
            related_items.append(f"{len(customer.attendance_records)} attendance record(s)")
        if has_history_events:
            related_items.append(f"{len(customer.history_events)} history event(s)")
        
        related_text = ", ".join(related_items)
        flash(f'Cannot delete customer {customer.first_name} {customer.last_name}. Customer has {related_text}. Please remove or transfer these items first.', 'error')
        return redirect(url_for('customers'))
    
    try:
        # Log the deletion activity
        log_activity(
            entity_type='customer',
            entity_id=customer.id,
            action='deleted',
            user_id=current_user.id,
            description=f'Customer {customer.first_name} {customer.last_name} deleted',
            old_values=json.dumps({
                'first_name': customer.first_name,
                'last_name': customer.last_name,
                'phone': customer.phone,
                'phone2': customer.phone2,
                'age': customer.age,
                'status': customer.status,
                'created_at': customer.created_at.isoformat() if customer.created_at else None
            })
        )
        
        # Store customer name for flash message before deletion
        customer_name = f"{customer.first_name} {customer.last_name}"
        
        # Delete the customer
        db.session.delete(customer)
        db.session.commit()
        
        flash(f'Customer {customer_name} has been deleted successfully', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting customer: {str(e)}', 'error')
    
    return redirect(url_for('customers'))

@app.route('/customers/campaign', methods=['POST'])
@login_required
def send_whatsapp_campaign():
    selected_customers = request.form.getlist('selected_customers')
    message_template = request.form.get('message_template', '')
    
    if not selected_customers:
        flash('Please select at least one customer', 'error')
        return redirect(url_for('customers'))
    
    if not message_template:
        flash('Please enter a message template', 'error')
        return redirect(url_for('customers'))
    
    # Get selected customers
    if current_user.role == 'instructor':
        customers = Customer.query.filter(
            Customer.id.in_(selected_customers),
            Customer.assigned_instructors.contains(current_user)
        ).all()
    else:
        customers = Customer.query.filter(Customer.id.in_(selected_customers)).all()
    
    # Generate WhatsApp links
    whatsapp_links = []
    for customer in customers:
        if customer.phone:
            # Replace template variables
            personalized_message = message_template.replace('{name}', f"{customer.first_name} {customer.last_name}")
            personalized_message = personalized_message.replace('{first_name}', customer.first_name)
            personalized_message = personalized_message.replace('{last_name}', customer.last_name)
            
            # Clean phone number (remove spaces, dashes, etc.)
            clean_phone = ''.join(filter(str.isdigit, customer.phone))
            
            # Create WhatsApp Web link
            whatsapp_url = f"https://wa.me/{clean_phone}?text={personalized_message}"
            whatsapp_links.append({
                'customer': customer,
                'url': whatsapp_url,
                'message': personalized_message
            })
    
    # For now, we'll return the links to the user to open manually
    # In a real application, you might want to integrate with WhatsApp Business API
    return render_template('campaign_results.html', 
                         whatsapp_links=whatsapp_links, 
                         message_template=message_template)

@app.route('/customers/<int:customer_id>')
@login_required
def customer_detail(customer_id):
    # Check if instructor can access this customer (must be in their groups)
    if current_user.role == 'instructor':
        # Check if customer is in any of instructor's groups
        instructor_groups = Group.query.filter_by(instructor_id=current_user.id).all()
        group_ids = [group.id for group in instructor_groups]
        
        if group_ids:
            customer_in_groups = GroupMember.query.filter(
                GroupMember.group_id.in_(group_ids),
                GroupMember.customer_id == customer_id,
                GroupMember.status == 'active'
            ).first()
            
            if not customer_in_groups:
                flash('Access denied. This student is not in your groups.', 'error')
                return redirect(url_for('customers'))
        else:
            flash('Access denied', 'error')
            return redirect(url_for('dashboard'))
    
    customer = Customer.query.get_or_404(customer_id)
    
    # Check permissions
    if current_user.role == 'instructor' and current_user not in customer.assigned_instructors:
        flash('Access denied', 'error')
        return redirect(url_for('customers'))
    
    # Get basic data
    notes = Note.query.filter_by(customer_id=customer_id).order_by(Note.created_at.desc()).all()
    tickets = Ticket.query.filter_by(customer_id=customer_id).order_by(Ticket.created_at.desc()).all()
    sessions = Session.query.filter_by(customer_id=customer_id).order_by(Session.scheduled_date.desc()).all()
    
    # Get group memberships and performance
    group_memberships = db.session.query(GroupMember, Group).join(Group).filter(
        GroupMember.customer_id == customer_id
    ).order_by(GroupMember.joined_date.desc()).all()
    
    # Get attendance records
    attendance_records = db.session.query(GroupAttendance, GroupSession, Group).join(
        GroupSession, GroupAttendance.group_session_id == GroupSession.id
    ).join(Group, GroupSession.group_id == Group.id).filter(
        GroupAttendance.customer_id == customer_id
    ).order_by(GroupSession.session_date.desc()).all()
    
    # Calculate attendance statistics
    total_sessions = len(attendance_records)
    present_sessions = len([r for r in attendance_records if r[0].status == 'present'])
    attendance_rate = (present_sessions / total_sessions * 100) if total_sessions > 0 else 0
    
    # Get performance records
    performance_records = Performance.query.filter_by(customer_id=customer_id).order_by(
        Performance.assessment_date.desc()
    ).all()
    
    # Calculate performance statistics
    avg_score = sum([p.score for p in performance_records if p.score]) / len(performance_records) if performance_records else None
    
    # Get communications
    communications = Communication.query.filter_by(customer_id=customer_id).order_by(
        Communication.created_at.desc()
    ).limit(10).all()
    
    # Get customer history
    customer_history = CustomerHistory.query.filter_by(customer_id=customer_id).order_by(
        CustomerHistory.created_at.desc()
    ).limit(20).all()
    
    # Group-specific performance
    group_performance = {}
    for membership, group in group_memberships:
        group_perfs = [p for p in performance_records if p.group_id == group.id]
        group_attendance = [a for a in attendance_records if a[2].id == group.id]
        
        group_present = len([a for a in group_attendance if a[0].status == 'present'])
        group_total = len(group_attendance)
        group_attendance_rate = (group_present / group_total * 100) if group_total > 0 else 0
        
        group_avg_score = sum([p.score for p in group_perfs if p.score]) / len(group_perfs) if group_perfs else None
        
        group_performance[group.id] = {
            'group': group,
            'membership': membership,
            'attendance_rate': group_attendance_rate,
            'total_sessions': group_total,
            'present_sessions': group_present,
            'avg_score': group_avg_score,
            'recent_performances': group_perfs[:3]
        }
    
    # Overall customer statistics
    customer_stats = {
        'total_groups': len([m for m, g in group_memberships if m.status == 'active']),
        'total_sessions_attended': present_sessions,
        'total_sessions': total_sessions,
        'attendance_rate': round(attendance_rate, 1),
        'avg_performance': round(avg_score, 1) if avg_score else None,
        'total_assessments': len(performance_records),
        'total_tickets': len(tickets),
        'open_tickets': len([t for t in tickets if t.status in ['open', 'in_progress']]),
        'total_communications': len(communications),
        'customer_since_days': (datetime.now().date() - customer.created_at.date()).days
    }
    
    # Monthly activity for charts
    monthly_activity = {}
    
    # Group sessions by month
    for attendance, session, group in attendance_records:
        month_key = session.session_date.strftime('%Y-%m')
        if month_key not in monthly_activity:
            monthly_activity[month_key] = {'sessions': 0, 'assessments': 0}
        monthly_activity[month_key]['sessions'] += 1
    
    # Assessments by month
    for performance in performance_records:
        month_key = performance.assessment_date.strftime('%Y-%m')
        if month_key not in monthly_activity:
            monthly_activity[month_key] = {'sessions': 0, 'assessments': 0}
        monthly_activity[month_key]['assessments'] += 1
    
    return render_template('customer_detail.html', 
                         customer=customer, 
                         notes=notes,
                         tickets=tickets, 
                         sessions=sessions,
                         group_memberships=group_memberships,
                         attendance_records=attendance_records[:10],  # Recent 10
                         performance_records=performance_records[:10],  # Recent 10
                         communications=communications,
                         customer_history=customer_history,
                         group_performance=group_performance,
                         customer_stats=customer_stats,
                         monthly_activity=monthly_activity)

# Excel Import Routes
@app.route('/customers/import-template')
@login_required
def download_customer_template():
    """Download Excel template for customer import"""
    # Create sample data for the template
    template_data = [
        {
            'first_name': 'محمد',
            'last_name': 'أحمد',
            'phone': '+966501234567',
            'phone2': '+966501234568',
            'age': 25,
            'status': 'active',
            'assigned_instructor_email': 'instructor@example.com',
            'initial_notes': 'ملاحظات أولية للعميل'
        },
        {
            'first_name': 'فاطمة',
            'last_name': 'سالم',
            'phone': '+966507654321',
            'phone2': '',
            'age': 30,
            'status': 'active',
            'assigned_instructor_email': '',
            'initial_notes': 'عميل مهتم بدورات البرمجة'
        },
        {
            'first_name': 'أحمد',
            'last_name': 'محمود',
            'phone': '+966509876543',
            'phone2': '+966509876544',
            'age': 22,
            'status': 'needs_follow_up',
            'assigned_instructor_email': 'instructor@example.com',
            'initial_notes': 'يحتاج متابعة خاصة'
        }
    ]
    
    # Create workbook
    wb = Workbook()
    
    # Create main template sheet
    ws = wb.active
    ws.title = 'Customers Template'
    
    # Add headers
    headers = list(template_data[0].keys())
    ws.append(headers)
    
    # Add sample data
    for row_data in template_data:
        row = [row_data.get(col, '') for col in headers]
        ws.append(row)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add instructions sheet
    instructions_sheet = wb.create_sheet(title='Instructions')
    instructions_data = [
        {
            'Column Name': 'first_name',
            'Required': 'Yes',
            'Description': 'الاسم الأول للعميل',
            'Example': 'محمد'
        },
        {
            'Column Name': 'last_name',
            'Required': 'Yes',
            'Description': 'اسم العائلة للعميل',
            'Example': 'أحمد'
        },
        {
            'Column Name': 'email',
            'Required': 'Yes',
            'Description': 'البريد الإلكتروني (يجب أن يكون فريد)',
            'Example': 'mohammed@example.com'
        },
        {
            'Column Name': 'phone',
            'Required': 'No',
            'Description': 'رقم الهاتف (مع رمز الدولة)',
            'Example': '+966501234567'
        },
        {
            'Column Name': 'status',
            'Required': 'No',
            'Description': 'حالة العميل: active, inactive, needs_follow_up, no_show',
            'Example': 'active'
        },
        {
            'Column Name': 'assigned_instructor_email',
            'Required': 'No',
            'Description': 'البريد الإلكتروني للمدرب المسؤول (اختياري)',
            'Example': 'instructor@example.com'
        },
        {
            'Column Name': 'initial_notes',
            'Required': 'No',
            'Description': 'ملاحظات أولية عن العميل',
            'Example': 'عميل مهتم بدورات البرمجة'
        }
    ]
    
    # Add instructions headers
    instructions_headers = list(instructions_data[0].keys())
    instructions_sheet.append(instructions_headers)
    
    # Add instructions data
    for row_data in instructions_data:
        row = [row_data.get(col, '') for col in instructions_headers]
        instructions_sheet.append(row)
    
    # Style instructions headers
    for cell in instructions_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths for instructions
    for column in instructions_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        instructions_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=customer_import_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    return response

@app.route('/customers/import', methods=['GET', 'POST'])
@login_required
def import_customers():
    """Import customers from Excel file"""
    if request.method == 'GET':
        return render_template('import_customers.html')
    
    # Handle file upload
    if 'excel_file' not in request.files:
        flash('يرجى اختيار ملف Excel', 'error')
        return redirect(request.url)
    
    file = request.files['excel_file']
    if file.filename == '':
        flash('لم يتم اختيار أي ملف', 'error')
        return redirect(request.url)
    
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        flash('يرجى رفع ملف Excel فقط (.xlsx أو .xls)', 'error')
        return redirect(request.url)
    
    try:
        # Save uploaded file temporarily
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            
            # Read Excel file using helper function
            data = read_excel_file(tmp_file.name)
            
            # Clean up temp file
            import os
            os.unlink(tmp_file.name)
        
        # Get column names from first row (if any data exists)
        if not data:
            flash('الملف فارغ أو لا يحتوي على بيانات', 'error')
            return redirect(request.url)
            
        found_columns = list(data[0].keys()) if data else []
        print(f"Found columns in Excel: {found_columns}")
        
        # Clean column names (remove extra spaces, etc.)
        cleaned_data = []
        for row in data:
            cleaned_row = {}
            for key, value in row.items():
                clean_key = key.strip() if key else ''
                cleaned_row[clean_key] = value
            cleaned_data.append(cleaned_row)
        data = cleaned_data
        
        # Update found_columns after cleaning
        found_columns = list(data[0].keys()) if data else []
        
        # Validate required columns
        required_columns = ['first_name', 'last_name']
        missing_columns = [col for col in required_columns if col not in found_columns]
        
        if missing_columns:
            flash(f'الأعمدة التالية مطلوبة ومفقودة: {", ".join(missing_columns)}', 'error')
            flash(f'الأعمدة الموجودة في الملف: {", ".join(found_columns)}', 'info')
            flash('يرجى التأكد من أن أسماء الأعمدة مطابقة تماماً: first_name, last_name', 'info')
            return redirect(request.url)
        
        # Import statistics
        imported_count = 0
        skipped_count = 0
        error_messages = []
        
        for index, row in enumerate(data):
            try:
                # Skip empty rows
                if is_empty_value(row.get('first_name')) or is_empty_value(row.get('last_name')):
                    skipped_count += 1
                    continue
                
                # Find assigned instructor if specified
                assigned_instructor = None
                if not is_empty_value(row.get('assigned_instructor_email')):
                    assigned_instructor = User.query.filter_by(
                        email=row['assigned_instructor_email'],
                        role='instructor'
                    ).first()
                    if not assigned_instructor:
                        error_messages.append(f'الصف {index + 2}: المدرب بالبريد الإلكتروني {row["assigned_instructor_email"]} غير موجود')
                
                # Validate status
                valid_statuses = ['active', 'inactive', 'needs_follow_up', 'no_show']
                status = row.get('status', 'active')
                if is_empty_value(status):
                    status = 'active'
                if status not in valid_statuses:
                    status = 'active'
                
                # Handle age conversion
                age = None
                if not is_empty_value(row.get('age')):
                    try:
                        age_value = row.get('age')
                        if isinstance(age_value, (int, float)) and age_value > 0:
                            age = int(age_value)
                    except (ValueError, TypeError):
                        pass  # Keep age as None if conversion fails
                
                # Create customer
                customer = Customer(
                    first_name=safe_str(row['first_name']),
                    last_name=safe_str(row['last_name']),
                    phone=safe_str(row.get('phone', '')),
                    phone2=safe_str(row.get('phone2', '')),
                    age=age,
                    status=status,
                    initial_notes=safe_str(row.get('initial_notes', ''))
                )
                
                db.session.add(customer)
                db.session.flush()  # Get the customer ID
                
                # Assign instructor if found
                if assigned_instructor:
                    customer.assigned_instructors.append(assigned_instructor)
                
                imported_count += 1
                
            except Exception as e:
                error_messages.append(f'الصف {index + 2}: خطأ في الاستيراد - {str(e)}')
                skipped_count += 1
        
        # Commit changes
        try:
            db.session.commit()
            
            # Success message
            success_msg = f'تم استيراد {imported_count} عميل بنجاح'
            if skipped_count > 0:
                success_msg += f', تم تخطي {skipped_count} صف'
            flash(success_msg, 'success')
            
            # Show errors if any
            if error_messages:
                for error in error_messages[:10]:  # Show first 10 errors
                    flash(error, 'warning')
                if len(error_messages) > 10:
                    flash(f'وهناك {len(error_messages) - 10} أخطاء إضافية...', 'warning')
            
            return redirect(url_for('customers'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'حدث خطأ أثناء حفظ البيانات: {str(e)}', 'error')
            return redirect(request.url)
            
    except Exception as e:
        flash(f'خطأ في قراءة ملف Excel: {str(e)}', 'error')
        return redirect(request.url)

# Ticket Management Routes
@app.route('/tickets')
@login_required
def tickets():
    # Define sort order for ticket status (lower number = higher priority/top position)
    status_priority = {
        'urgent': 0,
        'open': 1,
        'in_progress': 2,
        'resolved': 3,
        'closed': 4
    }
    
    if current_user.role == 'customer_service':
        tickets_query = Ticket.query.filter_by(assigned_to_id=current_user.id)
    elif current_user.role == 'instructor':
        # Instructors see tickets for their students only
        instructor_groups = Group.query.filter_by(instructor_id=current_user.id).all()
        group_ids = [group.id for group in instructor_groups]
        
        if group_ids:
            # Get customer IDs from instructor's groups
            customer_ids = db.session.query(GroupMember.customer_id).filter(
                GroupMember.group_id.in_(group_ids),
                GroupMember.status == 'active'
            ).distinct().all()
            customer_ids = [cid[0] for cid in customer_ids]
            
            if customer_ids:
                tickets_query = Ticket.query.filter(Ticket.customer_id.in_(customer_ids))
            else:
                tickets_query = Ticket.query.filter(Ticket.id == -1)  # No tickets
        else:
            tickets_query = Ticket.query.filter(Ticket.id == -1)  # No tickets
    else:
        tickets_query = Ticket.query
    
    # Get all tickets and sort them
    tickets = tickets_query.all()
    
    # Sort tickets by: 1) Status priority 2) Priority level 3) Creation date (newest first)
    priority_order = {'urgent': 0, 'high': 1, 'medium': 2, 'low': 3}
    
    tickets.sort(key=lambda t: (
        status_priority.get(t.status, 5),  # Status priority (resolved/closed go to bottom)
        priority_order.get(t.priority, 4),  # Priority level
        -t.id  # Newer tickets first (negative for descending order)
    ))
    
    return render_template('tickets.html', tickets=tickets)

@app.route('/tickets/add', methods=['GET', 'POST'])
@login_required
def add_ticket():
    if request.method == 'POST':
        # Check if instructor is trying to create ticket for their student
        if current_user.role == 'instructor':
            customer_id = int(request.form['customer_id'])
            instructor_groups = Group.query.filter_by(instructor_id=current_user.id).all()
            group_ids = [group.id for group in instructor_groups]
            
            if group_ids:
                customer_in_groups = GroupMember.query.filter(
                    GroupMember.group_id.in_(group_ids),
                    GroupMember.customer_id == customer_id,
                    GroupMember.status == 'active'
                ).first()
                
                if not customer_in_groups:
                    flash('Access denied. You can only create tickets for your students.', 'error')
                    return redirect(url_for('add_ticket'))
            else:
                flash('Access denied', 'error')
                return redirect(url_for('tickets'))
        
        ticket = Ticket(
            title=request.form['title'],
            description=request.form['description'],
            customer_id=request.form['customer_id'],
            assigned_to_id=request.form.get('assigned_to_id') or None,
            priority=request.form['priority']
        )
        db.session.add(ticket)
        db.session.commit()
        flash('Ticket created successfully', 'success')
        return redirect(url_for('tickets'))
    
    # Get customers based on user role
    if current_user.role == 'instructor':
        # Instructors see only their students
        instructor_groups = Group.query.filter_by(instructor_id=current_user.id).all()
        group_ids = [group.id for group in instructor_groups]
        
        if group_ids:
            customer_ids = db.session.query(GroupMember.customer_id).filter(
                GroupMember.group_id.in_(group_ids),
                GroupMember.status == 'active'
            ).distinct().all()
            customer_ids = [cid[0] for cid in customer_ids]
            
            if customer_ids:
                customers = Customer.query.filter(Customer.id.in_(customer_ids)).all()
            else:
                customers = []
        else:
            customers = []
    else:
        customers = Customer.query.all()
    
    agents = User.query.filter_by(role='customer_service').all()
    return render_template('add_ticket.html', customers=customers, agents=agents)

@app.route('/tickets/<int:ticket_id>')
@login_required
def ticket_detail(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    
    # Check permissions
    if current_user.role == 'customer_service' and ticket.assigned_to_id != current_user.id:
        flash('Access denied. You can only view tickets assigned to you.', 'error')
        return redirect(url_for('tickets'))
    elif current_user.role == 'instructor':
        # Check if ticket is for instructor's student
        instructor_groups = Group.query.filter_by(instructor_id=current_user.id).all()
        group_ids = [group.id for group in instructor_groups]
        
        if group_ids:
            customer_in_groups = GroupMember.query.filter(
                GroupMember.group_id.in_(group_ids),
                GroupMember.customer_id == ticket.customer_id,
                GroupMember.status == 'active'
            ).first()
            
            if not customer_in_groups:
                flash('Access denied. This ticket is not for your students.', 'error')
                return redirect(url_for('tickets'))
        else:
            flash('Access denied', 'error')
            return redirect(url_for('tickets'))
    
    return render_template('ticket_detail.html', ticket=ticket)

@app.route('/tickets/<int:ticket_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_ticket(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    
    # Check permissions
    if current_user.role == 'customer_service' and ticket.assigned_to_id != current_user.id:
        flash('Access denied. You can only edit tickets assigned to you.', 'error')
        return redirect(url_for('tickets'))
    elif current_user.role == 'instructor':
        # Check if ticket is for instructor's student
        instructor_groups = Group.query.filter_by(instructor_id=current_user.id).all()
        group_ids = [group.id for group in instructor_groups]
        
        if group_ids:
            customer_in_groups = GroupMember.query.filter(
                GroupMember.group_id.in_(group_ids),
                GroupMember.customer_id == ticket.customer_id,
                GroupMember.status == 'active'
            ).first()
            
            if not customer_in_groups:
                flash('Access denied. This ticket is not for your students.', 'error')
                return redirect(url_for('tickets'))
        else:
            flash('Access denied', 'error')
            return redirect(url_for('tickets'))
    
    if request.method == 'POST':
        ticket.title = request.form['title']
        ticket.description = request.form['description']
        ticket.priority = request.form['priority']
        ticket.status = request.form['status']
        
        # Only admins can reassign tickets
        if current_user.role == 'admin':
            ticket.assigned_to_id = request.form.get('assigned_to_id') or None
        
        # Update resolution info if status is resolved
        if ticket.status == 'resolved' and not ticket.resolved_at:
            ticket.resolved_at = datetime.utcnow()
            ticket.resolution_notes = request.form.get('resolution_notes', '')
        elif ticket.status != 'resolved':
            ticket.resolved_at = None
            ticket.resolution_notes = None
        
        db.session.commit()
        flash('Ticket updated successfully', 'success')
        return redirect(url_for('ticket_detail', ticket_id=ticket.id))
    
    # Get customers based on user role
    if current_user.role == 'instructor':
        # Instructors see only their students
        instructor_groups = Group.query.filter_by(instructor_id=current_user.id).all()
        group_ids = [group.id for group in instructor_groups]
        
        if group_ids:
            customer_ids = db.session.query(GroupMember.customer_id).filter(
                GroupMember.group_id.in_(group_ids),
                GroupMember.status == 'active'
            ).distinct().all()
            customer_ids = [cid[0] for cid in customer_ids]
            
            if customer_ids:
                customers = Customer.query.filter(Customer.id.in_(customer_ids)).all()
            else:
                customers = []
        else:
            customers = []
    else:
        customers = Customer.query.all()
    
    agents = User.query.filter_by(role='customer_service').all()
    return render_template('edit_ticket.html', ticket=ticket, customers=customers, agents=agents)

@app.route('/tickets/<int:ticket_id>/resolve', methods=['POST'])
@login_required
def resolve_ticket(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    
    # Check permissions
    if current_user.role == 'customer_service' and ticket.assigned_to_id != current_user.id:
        flash('Access denied. You can only resolve tickets assigned to you.', 'error')
        return redirect(url_for('tickets'))
    elif current_user.role == 'instructor':
        # Check if ticket is for instructor's student
        instructor_groups = Group.query.filter_by(instructor_id=current_user.id).all()
        group_ids = [group.id for group in instructor_groups]
        
        if group_ids:
            customer_in_groups = GroupMember.query.filter(
                GroupMember.group_id.in_(group_ids),
                GroupMember.customer_id == ticket.customer_id,
                GroupMember.status == 'active'
            ).first()
            
            if not customer_in_groups:
                flash('Access denied. This ticket is not for your students.', 'error')
                return redirect(url_for('tickets'))
        else:
            flash('Access denied', 'error')
            return redirect(url_for('tickets'))
    
    ticket.status = 'resolved'
    ticket.resolved_at = datetime.utcnow()
    ticket.resolution_notes = request.form.get('resolution_notes', '')
    
    db.session.commit()
    flash('Ticket marked as resolved', 'success')
    return redirect(url_for('tickets'))

# Session/Schedule Management Routes
@app.route('/sessions')
@login_required
def sessions():
    # Get filter parameters
    status_filter = request.args.get('status_filter')
    date_filter = request.args.get('date_filter')
    
    if current_user.role == 'instructor':
        # Instructors see only their own sessions
        query = Session.query.filter_by(instructor_id=current_user.id)
    else:
        # Admin and customer service see all sessions
        query = Session.query
    
    # Apply filters
    if status_filter and status_filter != 'all':
        query = query.filter(Session.status == status_filter)
    
    if date_filter:
        if date_filter == 'today':
            today = datetime.now().date()
            query = query.filter(Session.scheduled_date >= today, 
                               Session.scheduled_date < today + timedelta(days=1))
        elif date_filter == 'week':
            today = datetime.now().date()
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=7)
            query = query.filter(Session.scheduled_date >= week_start, 
                               Session.scheduled_date < week_end)
        elif date_filter == 'month':
            today = datetime.now().date()
            month_start = today.replace(day=1)
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year + 1, month=1)
            else:
                month_end = month_start.replace(month=month_start.month + 1)
            query = query.filter(Session.scheduled_date >= month_start, 
                               Session.scheduled_date < month_end)
    
    sessions = query.order_by(Session.scheduled_date.desc()).all()
    
    # Get statistics for display
    stats = {
        'total': len(sessions),
        'scheduled': len([s for s in sessions if s.status == 'scheduled']),
        'completed': len([s for s in sessions if s.status == 'completed']),
        'no_show': len([s for s in sessions if s.status == 'no_show']),
        'cancelled': len([s for s in sessions if s.status == 'cancelled'])
    }
    
    return render_template('sessions.html', sessions=sessions, stats=stats, 
                         status_filter=status_filter, date_filter=date_filter)

@app.route('/sessions/add', methods=['GET', 'POST'])
@login_required
def add_session():
    if request.method == 'POST':
        # Get multiple customer IDs from form
        customer_ids = request.form.getlist('customer_ids')
        
        if not customer_ids:
            flash('Please select at least one customer', 'error')
            customers = Customer.query.all()
            instructors = User.query.filter_by(role='instructor').all()
            return render_template('add_session.html', customers=customers, instructors=instructors)
        
        # Get common session data
        instructor_id = request.form['instructor_id']
        scheduled_date = datetime.strptime(request.form['scheduled_date'], '%Y-%m-%dT%H:%M')
        duration = int(request.form['duration'])
        notes = request.form.get('notes', '')
        
        # Create a session for each selected customer
        sessions_created = 0
        for customer_id in customer_ids:
            try:
                session_obj = Session(
                    customer_id=int(customer_id),
                    instructor_id=instructor_id,
                    course_id=None,  # No longer using courses
                    scheduled_date=scheduled_date,
                    duration=duration,
                    notes=notes
                )
                db.session.add(session_obj)
                sessions_created += 1
                
                # Log the session creation for each customer
                log_customer_event(
                    customer_id=int(customer_id),
                    event_type='session_scheduled',
                    description=f'Session scheduled for {scheduled_date.strftime("%Y-%m-%d %H:%M")}',
                    event_data=f'{{"duration": {duration}, "instructor_id": {instructor_id}}}',
                    created_by_id=current_user.id
                )
                
            except Exception as e:
                # If there's an error with one customer, continue with others
                continue
        
        if sessions_created > 0:
            db.session.commit()
            if sessions_created == 1:
                flash('Session scheduled successfully', 'success')
            else:
                flash(f'{sessions_created} sessions scheduled successfully', 'success')
        else:
            flash('No sessions could be created. Please try again.', 'error')
            
        return redirect(url_for('sessions'))
    
    customers = Customer.query.all()
    instructors = User.query.filter_by(role='instructor').all()
    return render_template('add_session.html', customers=customers, instructors=instructors)

@app.route('/sessions/<int:session_id>')
@login_required
def session_detail(session_id):
    session = Session.query.get_or_404(session_id)
    
    # Check if user has permission to view this session
    if current_user.role == 'instructor' and session.instructor_id != current_user.id:
        flash('You do not have permission to view this session', 'error')
        return redirect(url_for('sessions'))
    
    return render_template('session_detail.html', session=session)

@app.route('/sessions/<int:session_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_session(session_id):
    session = Session.query.get_or_404(session_id)
    
    # Check if user has permission to edit this session
    if current_user.role == 'instructor' and session.instructor_id != current_user.id:
        flash('You do not have permission to edit this session', 'error')
        return redirect(url_for('sessions'))
    
    if request.method == 'POST':
        # Get form data
        customer_id = request.form['customer_id']
        instructor_id = request.form['instructor_id']
        scheduled_date = datetime.strptime(request.form['scheduled_date'], '%Y-%m-%dT%H:%M')
        duration = int(request.form['duration'])
        notes = request.form.get('notes', '')
        
        # Store old values for logging
        old_values = {
            'customer_id': session.customer_id,
            'instructor_id': session.instructor_id,
            'scheduled_date': session.scheduled_date.isoformat(),
            'duration': session.duration,
            'notes': session.notes
        }
        
        # Update session
        session.customer_id = customer_id
        session.instructor_id = instructor_id
        session.scheduled_date = scheduled_date
        session.duration = duration
        session.notes = notes
        
        # Log the activity
        log_activity(
            entity_type='session',
            entity_id=session.id,
            action='updated',
            user_id=current_user.id,
            description=f'Session #{session.id} updated',
            old_values=old_values,
            new_values={
                'customer_id': int(customer_id),
                'instructor_id': int(instructor_id),
                'scheduled_date': scheduled_date.isoformat(),
                'duration': duration,
                'notes': notes
            }
        )
        
        # Log customer event
        log_customer_event(
            customer_id=int(customer_id),
            event_type='session_updated',
            description=f'Session scheduled for {scheduled_date.strftime("%Y-%m-%d %H:%M")} was updated',
            event_data={'session_id': session.id, 'duration': duration},
            created_by_id=current_user.id
        )
        
        db.session.commit()
        flash('Session updated successfully', 'success')
        return redirect(url_for('sessions'))
    
    customers = Customer.query.all()
    instructors = User.query.filter_by(role='instructor').all()
    return render_template('edit_session.html', session=session, customers=customers, instructors=instructors)

@app.route('/sessions/<int:session_id>/status', methods=['POST'])
@login_required
def update_session_status(session_id):
    session = Session.query.get_or_404(session_id)
    
    # Check if user has permission to update this session
    if current_user.role == 'instructor' and session.instructor_id != current_user.id:
        return jsonify({'success': False, 'message': 'Permission denied'}), 403
    
    new_status = request.json.get('status')
    if new_status not in ['completed', 'no_show', 'cancelled', 'scheduled']:
        return jsonify({'success': False, 'message': 'Invalid status'}), 400
    
    old_status = session.status
    session.status = new_status
    
    # Log the activity
    log_activity(
        entity_type='session',
        entity_id=session.id,
        action='status_updated',
        user_id=current_user.id,
        description=f'Session #{session.id} status changed from {old_status} to {new_status}',
        old_values={'status': old_status},
        new_values={'status': new_status}
    )
    
    # Log customer event
    event_description = f'Session status changed to {new_status.replace("_", " ").title()}'
    log_customer_event(
        customer_id=session.customer_id,
        event_type='session_status_updated',
        description=event_description,
        event_data={'session_id': session.id, 'old_status': old_status, 'new_status': new_status},
        created_by_id=current_user.id
    )
    
    db.session.commit()
    
    return jsonify({
        'success': True, 
        'message': f'Session marked as {new_status.replace("_", " ")}',
        'new_status': new_status
    })

# Attendance Reports and Statistics Routes
@app.route('/admin/attendance-reports')
@login_required
def attendance_reports():
    """Attendance reports and statistics for admin and customer service"""
    if current_user.role not in ['admin', 'customer_service']:
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    # Get filter parameters
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    instructor_filter = request.args.get('instructor_filter')
    group_filter = request.args.get('group_filter')
    
    # Build base queries
    sessions_query = Session.query
    group_sessions_query = GroupSession.query
    
    # Apply date filters
    if date_from:
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        sessions_query = sessions_query.filter(Session.scheduled_date >= date_from_obj)
        group_sessions_query = group_sessions_query.filter(GroupSession.session_date >= date_from_obj)
    
    if date_to:
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        sessions_query = sessions_query.filter(Session.scheduled_date <= date_to_obj)
        group_sessions_query = group_sessions_query.filter(GroupSession.session_date <= date_to_obj)
    
    # Apply instructor filter
    if instructor_filter:
        sessions_query = sessions_query.filter(Session.instructor_id == instructor_filter)
        group_sessions_query = group_sessions_query.join(Group).filter(Group.instructor_id == instructor_filter)
    
    # Apply group filter for group sessions
    if group_filter:
        group_sessions_query = group_sessions_query.filter(GroupSession.group_id == group_filter)
    
    # Get individual sessions data
    individual_sessions = sessions_query.all()
    
    # Get group sessions data
    group_sessions = group_sessions_query.all()
    
    # Calculate individual session statistics
    individual_stats = {
        'total_sessions': len(individual_sessions),
        'completed': len([s for s in individual_sessions if s.status == 'completed']),
        'no_show': len([s for s in individual_sessions if s.status == 'no_show']),
        'cancelled': len([s for s in individual_sessions if s.status == 'cancelled']),
        'scheduled': len([s for s in individual_sessions if s.status == 'scheduled']),
    }
    
    # Calculate attendance rate for individual sessions
    if individual_stats['total_sessions'] > 0:
        individual_stats['attendance_rate'] = round(
            (individual_stats['completed'] / individual_stats['total_sessions']) * 100, 2
        )
    else:
        individual_stats['attendance_rate'] = 0
    
    # Calculate group session statistics
    group_stats = {
        'total_sessions': len(group_sessions),
        'completed': len([s for s in group_sessions if s.status == 'completed']),
        'cancelled': len([s for s in group_sessions if s.status == 'cancelled']),
        'scheduled': len([s for s in group_sessions if s.status == 'scheduled']),
    }
    
    # Get detailed attendance for group sessions
    group_attendance_details = []
    for session in group_sessions:
        if session.status == 'completed':
            attendance_records = GroupAttendance.query.filter_by(group_session_id=session.id).all()
            total_members = len(attendance_records)
            present_count = len([a for a in attendance_records if a.status == 'present'])
            absent_count = len([a for a in attendance_records if a.status == 'absent'])
            late_count = len([a for a in attendance_records if a.status == 'late'])
            
            group_attendance_details.append({
                'session': session,
                'total_members': total_members,
                'present_count': present_count,
                'absent_count': absent_count,
                'late_count': late_count,
                'attendance_rate': round((present_count / total_members * 100), 2) if total_members > 0 else 0
            })
    
    # Get instructors and groups for filters
    instructors = User.query.filter_by(role='instructor', is_active=True).all()
    groups = Group.query.filter_by(status='active').all()
    
    return render_template('admin/attendance_reports.html',
                         individual_stats=individual_stats,
                         group_stats=group_stats,
                         group_attendance_details=group_attendance_details,
                         individual_sessions=individual_sessions,
                         instructors=instructors,
                         groups=groups,
                         date_from=date_from,
                         date_to=date_to,
                         instructor_filter=instructor_filter,
                         group_filter=group_filter)

@app.route('/instructor/quick-attendance')
@login_required
def instructor_quick_attendance():
    """Quick attendance interface for instructors"""
    if current_user.role != 'instructor':
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    # Get today's sessions for this instructor
    today = datetime.now().date()
    todays_sessions = Session.query.filter(
        Session.instructor_id == current_user.id,
        Session.scheduled_date >= today,
        Session.scheduled_date < today + timedelta(days=1),
        Session.status == 'scheduled'
    ).order_by(Session.scheduled_date).all()
    
    # Get today's group sessions for this instructor
    todays_group_sessions = GroupSession.query.join(Group).filter(
        Group.instructor_id == current_user.id,
        GroupSession.session_date == today,
        GroupSession.status == 'scheduled'
    ).order_by(GroupSession.start_time).all()
    
    # Format date in Arabic
    arabic_days = {
        'Monday': 'الاثنين',
        'Tuesday': 'الثلاثاء', 
        'Wednesday': 'الأربعاء',
        'Thursday': 'الخميس',
        'Friday': 'الجمعة',
        'Saturday': 'السبت',
        'Sunday': 'الأحد'
    }
    arabic_months = {
        'January': 'يناير',
        'February': 'فبراير',
        'March': 'مارس',
        'April': 'أبريل',
        'May': 'مايو',
        'June': 'يونيو',
        'July': 'يوليو',
        'August': 'أغسطس',
        'September': 'سبتمبر',
        'October': 'أكتوبر',
        'November': 'نوفمبر',
        'December': 'ديسمبر'
    }
    
    day_name = arabic_days.get(today.strftime('%A'), today.strftime('%A'))
    month_name = arabic_months.get(today.strftime('%B'), today.strftime('%B'))
    formatted_date = f"{day_name}، {today.day} {month_name} {today.year}"
    
    return render_template('instructor/quick_attendance.html',
                         todays_sessions=todays_sessions,
                         todays_group_sessions=todays_group_sessions,
                         current_date=today,
                         formatted_date=formatted_date)

@app.route('/groups/<int:group_id>/members/<int:customer_id>/attendance')
@login_required
def member_attendance_detail(group_id, customer_id):
    """Detailed attendance view for a specific member in a group"""
    group = Group.query.get_or_404(group_id)
    customer = Customer.query.get_or_404(customer_id)
    
    # Check permissions
    if current_user.role == 'instructor' and group.instructor_id != current_user.id:
        if current_user.role != 'admin' and current_user.role != 'customer_service':
            flash('Access denied', 'error')
            return redirect(url_for('groups'))
    
    # Verify the customer is a member of this group
    membership = GroupMember.query.filter_by(
        group_id=group_id, 
        customer_id=customer_id, 
        status='active'
    ).first()
    
    if not membership:
        flash('Customer is not an active member of this group', 'error')
        return redirect(url_for('group_detail', group_id=group_id))
    
    # Get all attendance records for this member in this group
    attendance_records = db.session.query(GroupAttendance, GroupSession).join(GroupSession).filter(
        GroupSession.group_id == group_id,
        GroupAttendance.customer_id == customer_id
    ).order_by(GroupSession.session_date.desc(), GroupSession.start_time.desc()).all()
    
    # Calculate detailed statistics
    total_sessions = len(attendance_records)
    present_count = len([a for a, s in attendance_records if a.status == 'present'])
    late_count = len([a for a, s in attendance_records if a.status == 'late'])
    absent_count = len([a for a, s in attendance_records if a.status == 'absent'])
    excused_count = len([a for a, s in attendance_records if a.status == 'excused'])
    
    attendance_rate = (present_count / total_sessions * 100) if total_sessions > 0 else 0
    
    # Calculate attendance streaks
    current_streak = 0
    longest_streak = 0
    temp_streak = 0
    
    for attendance, session in attendance_records:
        if attendance.status == 'present':
            if current_streak == 0:  # Start counting from most recent
                current_streak += 1
            temp_streak += 1
            longest_streak = max(longest_streak, temp_streak)
        else:
            if current_streak > 0:  # Stop counting recent streak
                current_streak = 0
            temp_streak = 0
    
    # Monthly attendance breakdown
    monthly_stats = {}
    for attendance, session in attendance_records:
        month_key = session.session_date.strftime('%Y-%m')
        if month_key not in monthly_stats:
            monthly_stats[month_key] = {
                'month_name': session.session_date.strftime('%B %Y'),
                'total': 0,
                'present': 0,
                'late': 0,
                'absent': 0,
                'excused': 0
            }
        
        monthly_stats[month_key]['total'] += 1
        monthly_stats[month_key][attendance.status] += 1
    
    # Calculate monthly attendance rates
    for month_data in monthly_stats.values():
        month_data['attendance_rate'] = round(
            (month_data['present'] / month_data['total'] * 100) if month_data['total'] > 0 else 0, 1
        )
    
    # Get recent performance records
    performance_records = Performance.query.filter_by(
        customer_id=customer_id,
        group_id=group_id
    ).order_by(Performance.assessment_date.desc()).limit(10).all()
    
    attendance_stats = {
        'total_sessions': total_sessions,
        'present_count': present_count,
        'late_count': late_count,
        'absent_count': absent_count,
        'excused_count': excused_count,
        'attendance_rate': round(attendance_rate, 1),
        'current_streak': current_streak,
        'longest_streak': longest_streak,
        'monthly_stats': monthly_stats
    }
    
    return render_template('group_member_attendance.html',
                         group=group,
                         customer=customer,
                         membership=membership,
                         attendance_records=attendance_records,
                         attendance_stats=attendance_stats,
                         performance_records=performance_records)

# API Routes for AJAX
@app.route('/api/dashboard_stats')
@login_required
def api_dashboard_stats():
    if current_user.role == 'admin':
        stats = {
            'total_users': User.query.count(),
            'total_customers': Customer.query.count(),
            'total_tickets': Ticket.query.count(),
            'open_tickets': Ticket.query.filter_by(status='open').count()
        }
    elif current_user.role == 'instructor':
        # Get today's sessions count (both individual and group sessions)
        today = datetime.now().date()
        individual_sessions_today = Session.query.filter_by(instructor_id=current_user.id).filter(
            Session.scheduled_date >= today,
            Session.scheduled_date < today + timedelta(days=1)
        ).count()
        
        group_sessions_today = GroupSession.query.join(Group).filter(
            Group.instructor_id == current_user.id,
            GroupSession.session_date == today
        ).count()
        
        # Get this week's sessions count
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        
        weekly_sessions = GroupSession.query.join(Group).filter(
            Group.instructor_id == current_user.id,
            GroupSession.session_date >= week_start,
            GroupSession.session_date <= week_end
        ).count()
        
        # Get active groups
        active_groups = Group.query.filter_by(instructor_id=current_user.id, status='active').count()
        
        # Get total students across all groups
        total_students = db.session.query(GroupMember).join(Group).filter(
            Group.instructor_id == current_user.id,
            GroupMember.status == 'active'
        ).count()
        
        # Get completed sessions this month
        month_start = today.replace(day=1)
        completed_sessions = Session.query.filter_by(instructor_id=current_user.id, status='completed').filter(
            Session.scheduled_date >= month_start
        ).count()
        
        stats = {
            'today_sessions': individual_sessions_today + group_sessions_today,
            'completed_sessions': completed_sessions,
            'active_groups': active_groups,
            'total_students': total_students,
            'weekly_sessions': weekly_sessions
        }
    elif current_user.role == 'customer_service':
        stats = {
            'my_tickets': Ticket.query.filter_by(assigned_to_id=current_user.id).count(),
            'open_tickets': Ticket.query.filter_by(assigned_to_id=current_user.id, status='open').count(),
            'resolved_today': Ticket.query.filter_by(assigned_to_id=current_user.id, status='resolved').filter(
                Ticket.resolved_at >= datetime.now().date()
            ).count()
        }
    
    return jsonify(stats)

# Course Management Routes
# Group Categories Management
@app.route('/admin/categories')
@admin_required
def admin_categories():
    categories = CourseCategory.query.all()
    return render_template('admin/categories.html', categories=categories)

@app.route('/admin/categories/add', methods=['POST'])
@admin_required
def admin_add_category():
    name = request.form.get('name')
    description = request.form.get('description')
    
    if name:
        category = CourseCategory(name=name, description=description)
        db.session.add(category)
        db.session.commit()
        flash('Category created successfully', 'success')
    else:
        flash('Category name is required', 'error')
    
    return redirect(url_for('admin_categories'))

@app.route('/admin/categories/<int:category_id>/edit', methods=['POST'])
@admin_required
def admin_edit_category(category_id):
    category = CourseCategory.query.get_or_404(category_id)
    
    name = request.form.get('name')
    description = request.form.get('description')
    
    if name:
        category.name = name
        category.description = description
        db.session.commit()
        flash('Category updated successfully', 'success')
    else:
        flash('Category name is required', 'error')
    
    return redirect(url_for('admin_categories'))

@app.route('/admin/categories/<int:category_id>/delete', methods=['POST'])
@admin_required
def admin_delete_category(category_id):
    category = CourseCategory.query.get_or_404(category_id)
    
    # Check if category has groups
    if category.category_groups:
        flash('Cannot delete category with existing groups', 'error')
    else:
        db.session.delete(category)
        db.session.commit()
        flash('Category deleted successfully', 'success')
    
    return redirect(url_for('admin_categories'))

# Report Generation Routes
@app.route('/admin/reports/customers/csv')
@admin_required
def export_customers_csv():
    # Get all customers with related data
    customers = Customer.query.all()
    
    # Prepare data for CSV
    data = []
    for customer in customers:
        data.append({
            'ID': customer.id,
            'First Name': customer.first_name,
            'Last Name': customer.last_name,
            'Phone': customer.phone or '',
            'Phone 2': customer.phone2 or '',
            'Age': customer.age or '',
            'Status': customer.status,
            'Assigned Instructors': ', '.join([f"{instructor.first_name} {instructor.last_name}" for instructor in customer.assigned_instructors]) if customer.assigned_instructors else 'Unassigned',
            'Created Date': customer.created_at.strftime('%Y-%m-%d'),
            'Notes': customer.initial_notes or '',
            'Total Sessions': len(customer.sessions),
            'Total Tickets': len(customer.tickets),
            'Active Group Memberships': len([m for m in customer.group_memberships if m.status == 'active'])
        })
    
    # Create CSV response using built-in csv module instead of pandas
    output = io.StringIO()
    if data:
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=customers_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

    return response

@app.route('/admin/reports/tickets/excel')
@admin_required
def export_tickets_excel():
    # Get all tickets with related data
    tickets = Ticket.query.all()
    
    # Prepare data for Excel
    data = []
    for ticket in tickets:
        data.append({
            'Ticket ID': ticket.id,
            'Title': ticket.title,
            'Description': ticket.description[:100] + '...' if len(ticket.description) > 100 else ticket.description,
            'Customer': f"{ticket.customer.first_name} {ticket.customer.last_name}",
            'Customer Phone': ticket.customer.phone or '',
            'Assigned To': f"{ticket.assigned_to.first_name} {ticket.assigned_to.last_name}" if ticket.assigned_to else 'Unassigned',
            'Status': ticket.status,
            'Priority': ticket.priority,
            'Created Date': ticket.created_at.strftime('%Y-%m-%d %H:%M'),
            'Resolved Date': ticket.resolved_at.strftime('%Y-%m-%d %H:%M') if ticket.resolved_at else 'Not Resolved',
            'Resolution Notes': ticket.resolution_notes or ''
        })
    
    # Create Excel file using openpyxl instead of pandas
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tickets Report'
    
    if data:
        # Add headers
        headers = list(data[0].keys())
        ws.append(headers)
        
        # Add data rows
        for row_data in data:
            row = [row_data.get(col, '') for col in headers]
            ws.append(row)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=tickets_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    return response

@app.route('/admin/reports/summary/pdf')
@admin_required
def export_summary_pdf():
    # Get statistics
    total_customers = Customer.query.count()
    total_groups = Group.query.count()
    total_tickets = Ticket.query.count()
    total_group_members = GroupMember.query.count()
    
    open_tickets = Ticket.query.filter_by(status='open').count()
    resolved_tickets = Ticket.query.filter_by(status='resolved').count()
    
    active_customers = Customer.query.filter_by(status='active').count()
    active_groups = Group.query.filter_by(status='active').count()
    
    # Create PDF in memory
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    
    # Container for PDF elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=12,
        textColor=colors.darkblue
    )
    
    # Title
    title = Paragraph("GENIO TECH CRM - Summary Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Report date
    date_para = Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %H:%M')}", styles['Normal'])
    elements.append(date_para)
    elements.append(Spacer(1, 20))
    
    # Overview Section
    overview_title = Paragraph("System Overview", heading_style)
    elements.append(overview_title)
    
    overview_data = [
        ['Metric', 'Count'],
        ['Total Customers', str(total_customers)],
        ['Active Customers', str(active_customers)],
        ['Total Groups', str(total_groups)],
        ['Active Groups', str(active_groups)],
        ['Total Group Members', str(total_group_members)],
        ['Total Tickets', str(total_tickets)],
        ['Open Tickets', str(open_tickets)],
        ['Resolved Tickets', str(resolved_tickets)]
    ]
    
    overview_table = Table(overview_data, colWidths=[3*inch, 2*inch])
    overview_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(overview_table)
    elements.append(Spacer(1, 20))
    
    # Customer Status Breakdown
    customer_title = Paragraph("Customer Status Breakdown", heading_style)
    elements.append(customer_title)
    
    customer_statuses = {}
    for status in ['active', 'inactive', 'needs_follow_up', 'no_show']:
        count = Customer.query.filter_by(status=status).count()
        customer_statuses[status.replace('_', ' ').title()] = count
    
    customer_data = [['Status', 'Count']]
    for status, count in customer_statuses.items():
        customer_data.append([status, str(count)])
    
    customer_table = Table(customer_data, colWidths=[3*inch, 2*inch])
    customer_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(customer_table)
    elements.append(Spacer(1, 20))
    
    # Recent Activity
    activity_title = Paragraph("Recent Activity (Last 30 Days)", heading_style)
    elements.append(activity_title)
    
    thirty_days_ago = datetime.now() - timedelta(days=30)
    recent_customers = Customer.query.filter(Customer.created_at >= thirty_days_ago).count()
    recent_tickets = Ticket.query.filter(Ticket.created_at >= thirty_days_ago).count()
    recent_group_members = GroupMember.query.filter(GroupMember.joined_date >= thirty_days_ago).count()
    
    activity_data = [
        ['Activity', 'Count'],
        ['New Customers', str(recent_customers)],
        ['New Tickets', str(recent_tickets)],
        ['New Group Members', str(recent_group_members)]
    ]
    
    activity_table = Table(activity_data, colWidths=[3*inch, 2*inch])
    activity_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(activity_table)
    
    # Footer
    elements.append(Spacer(1, 30))
    footer = Paragraph("Generated by GENIO TECH GENIOTECH", styles['Normal'])
    elements.append(footer)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=summary_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response

@app.route('/admin/reports/analytics')
@admin_required
def analytics_dashboard():
    # Generate analytics charts and data
    
    # Customer registration trend (last 12 months)
    months = []
    customer_counts = []
    for i in range(12):
        month_start = datetime.now().replace(day=1) - timedelta(days=30*i)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        count = Customer.query.filter(
            Customer.created_at >= month_start,
            Customer.created_at <= month_end
        ).count()
        months.append(month_start.strftime('%b %Y'))
        customer_counts.append(count)
    
    months.reverse()
    customer_counts.reverse()
    
    # Create customer trend chart
    plt.figure(figsize=(12, 6))
    plt.subplot(2, 2, 1)
    plt.plot(months, customer_counts, marker='o', linewidth=2, markersize=6)
    plt.title('Customer Registration Trend (12 Months)', fontsize=14, fontweight='bold')
    plt.xlabel('Month')
    plt.ylabel('New Customers')
    plt.xticks(rotation=45)
    plt.grid(True, alpha=0.3)
    
    # Ticket status distribution
    plt.subplot(2, 2, 2)
    ticket_statuses = ['open', 'in_progress', 'resolved', 'closed']
    ticket_counts = [Ticket.query.filter_by(status=status).count() for status in ticket_statuses]
    colors_pie = ['#ff9999', '#66b3ff', '#99ff99', '#ffcc99']
    
    # Check if we have any ticket data
    if sum(ticket_counts) > 0:
        # Filter out zero counts for better pie chart
        non_zero_data = [(status, count, color) for status, count, color in zip(ticket_statuses, ticket_counts, colors_pie) if count > 0]
        if non_zero_data:
            filtered_statuses, filtered_counts, filtered_colors = zip(*non_zero_data)
            plt.pie(filtered_counts, labels=filtered_statuses, autopct='%1.1f%%', colors=filtered_colors)
        else:
            plt.pie([1], labels=['No Data'], colors=['#cccccc'])
    else:
        plt.pie([1], labels=['No Tickets'], colors=['#cccccc'])
    plt.title('Ticket Status Distribution', fontsize=14, fontweight='bold')
    
    # Group membership stats
    plt.subplot(2, 2, 3)
    groups = Group.query.all()
    group_names = [group.name[:15] + '...' if len(group.name) > 15 else group.name for group in groups[:5]]
    member_counts = [len(group.members) for group in groups[:5]]
    plt.bar(group_names, member_counts, color='skyblue')
    plt.title('Top 5 Groups by Membership', fontsize=14, fontweight='bold')
    plt.xlabel('Group')
    plt.ylabel('Members')
    plt.xticks(rotation=45)
    
    # Customer status distribution
    plt.subplot(2, 2, 4)
    customer_statuses = ['active', 'inactive', 'needs_follow_up', 'no_show']
    customer_status_counts = [Customer.query.filter_by(status=status).count() for status in customer_statuses]
    plt.bar(customer_statuses, customer_status_counts, color=['green', 'gray', 'orange', 'red'])
    plt.title('Customer Status Distribution', fontsize=14, fontweight='bold')
    plt.xlabel('Status')
    plt.ylabel('Count')
    plt.xticks(rotation=45)
    
    plt.tight_layout()
    
    # Save chart to base64 string
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
    img_buffer.seek(0)
    
    # Convert to base64
    chart_data = base64.b64encode(img_buffer.getvalue()).decode()
    plt.close()
    
    # Calculate additional statistics
    stats = {
        'total_customers': Customer.query.count(),
        'total_groups': Group.query.count(),
        'total_tickets': Ticket.query.count(),
        'total_group_members': GroupMember.query.count(),
        'resolution_rate': (Ticket.query.filter_by(status='resolved').count() / max(Ticket.query.count(), 1)) * 100,
        'active_group_rate': (Group.query.filter_by(status='active').count() / max(Group.query.count(), 1)) * 100,
        'average_members_per_group': GroupMember.query.count() / max(Group.query.count(), 1)
    }
    
    return render_template('admin/analytics.html', chart_data=chart_data, stats=stats)

# Group Management Routes
@app.route('/groups')
@login_required
def groups():
    if current_user.role == 'instructor':
        groups = Group.query.filter_by(instructor_id=current_user.id).all()
    else:
        groups = Group.query.all()
    
    # Add member count to each group
    for group in groups:
        group.member_count = GroupMember.query.filter_by(group_id=group.id, status='active').count()
    
    return render_template('groups.html', groups=groups)

@app.route('/groups/weekly')
@login_required
def groups_weekly():
    """Show groups organized by day of the week"""
    if current_user.role == 'instructor':
        groups = Group.query.filter_by(instructor_id=current_user.id, status='active').all()
    else:
        groups = Group.query.filter_by(status='active').all()
    
    # Define days of the week
    days_of_week = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
    day_labels = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    
    # Get current day
    from datetime import datetime
    current_day = datetime.now().strftime('%A').lower()  # Get current day name in lowercase
    
    # Organize groups by day
    weekly_schedule = {}
    for i, day in enumerate(days_of_week):
        weekly_schedule[day] = {
            'label': day_labels[i],
            'groups': [],
            'is_today': day == current_day  # Mark if this is today
        }
    
    # Group schedules by day with group information
    for group in groups:
        for schedule in group.schedules:
            day = schedule.day_of_week.lower()
            if day in weekly_schedule:
                # Add member count to group
                group.member_count = GroupMember.query.filter_by(group_id=group.id, status='active').count()
                
                group_info = {
                    'group': group,
                    'schedule': schedule,
                    'member_count': group.member_count
                }
                weekly_schedule[day]['groups'].append(group_info)
    
    # Sort groups by start time for each day
    for day in weekly_schedule:
        weekly_schedule[day]['groups'].sort(key=lambda x: x['schedule'].start_time)
    
    return render_template('groups_weekly.html', weekly_schedule=weekly_schedule, days_of_week=days_of_week, current_day=current_day)

@app.route('/groups/add', methods=['GET', 'POST'])
@login_required
def add_group():
    # Prevent instructors from creating new groups
    if current_user.role == 'instructor':
        flash('Access denied. Instructors cannot create new groups.', 'error')
        return redirect(url_for('groups'))
    
    if request.method == 'POST':
        group = Group(
            name=request.form['name'],
            subject=request.form['subject'],
            description=request.form.get('description', ''),
            instructor_id=request.form['instructor_id'],
            category_id=request.form.get('category_id') if request.form.get('category_id') else None,
            max_students=int(request.form.get('max_students', 15)),
            start_date=datetime.strptime(request.form['start_date'], '%Y-%m-%d').date(),
            end_date=datetime.strptime(request.form['end_date'], '%Y-%m-%d').date() if request.form.get('end_date') else None,
            notes=request.form.get('notes', '')
        )
        
        db.session.add(group)
        db.session.flush()  # Get the group ID
        
        # Add schedules
        days = request.form.getlist('days[]')
        start_times = request.form.getlist('start_times[]')
        end_times = request.form.getlist('end_times[]')
        locations = request.form.getlist('locations[]')
        
        for i, day in enumerate(days):
            if day and i < len(start_times) and i < len(end_times):
                schedule = GroupSchedule(
                    group_id=group.id,
                    day_of_week=day,
                    start_time=datetime.strptime(start_times[i], '%H:%M').time(),
                    end_time=datetime.strptime(end_times[i], '%H:%M').time(),
                    location=locations[i] if i < len(locations) else ''
                )
                db.session.add(schedule)
        
        db.session.commit()
        flash('Group created successfully!', 'success')
        return redirect(url_for('groups'))
    
    instructors = User.query.filter_by(role='instructor', is_active=True).all()
    categories = CourseCategory.query.all()
    return render_template('add_group.html', instructors=instructors, categories=categories)

@app.route('/groups/<int:group_id>')
@login_required
def group_detail(group_id):
    group = Group.query.get_or_404(group_id)
    
    # Check permissions
    if current_user.role == 'instructor' and group.instructor_id != current_user.id:
        flash('Access denied.', 'error')
        return redirect(url_for('groups'))
    
    # Get group members with detailed info
    members = db.session.query(GroupMember, Customer).join(Customer).filter(
        GroupMember.group_id == group_id,
        GroupMember.status == 'active'
    ).all()
    
    # Get all members (including inactive) for history
    all_members = db.session.query(GroupMember, Customer).join(Customer).filter(
        GroupMember.group_id == group_id
    ).order_by(GroupMember.joined_date.desc()).all()
    
    # Get upcoming sessions
    upcoming_sessions = GroupSession.query.filter(
        GroupSession.group_id == group_id,
        GroupSession.session_date >= datetime.now().date(),
        GroupSession.status != 'cancelled'
    ).order_by(GroupSession.session_date, GroupSession.start_time).limit(10).all()
    
    # Get recent sessions with attendance
    recent_sessions = GroupSession.query.filter(
        GroupSession.group_id == group_id,
        GroupSession.session_date < datetime.now().date()
    ).order_by(GroupSession.session_date.desc(), GroupSession.start_time.desc()).limit(10).all()
    
    # Get all sessions for analytics
    all_sessions = GroupSession.query.filter_by(group_id=group_id).all()
    
    # Calculate analytics
    total_sessions = len(all_sessions)
    completed_sessions = len([s for s in all_sessions if s.status == 'completed'])
    cancelled_sessions = len([s for s in all_sessions if s.status == 'cancelled'])
    
    # Attendance analytics
    attendance_stats = {}
    total_possible_attendance = 0
    total_present = 0
    
    for session in all_sessions:
        if session.status == 'completed':
            session_attendance = GroupAttendance.query.filter_by(group_session_id=session.id).all()
            total_possible_attendance += len(session_attendance)
            total_present += len([a for a in session_attendance if a.status == 'present'])
    
    attendance_rate = (total_present / total_possible_attendance * 100) if total_possible_attendance > 0 else 0
    
    # Member performance
    member_performance = []
    for member, customer in members:
        # Get attendance for this member
        member_attendance = db.session.query(GroupAttendance).join(GroupSession).filter(
            GroupSession.group_id == group_id,
            GroupAttendance.customer_id == customer.id,
            GroupSession.status == 'completed'
        ).all()
        
        total_sessions_attended = len([a for a in member_attendance if a.status == 'present'])
        total_sessions_possible = len(member_attendance)
        attendance_percentage = (total_sessions_attended / total_sessions_possible * 100) if total_sessions_possible > 0 else 0
        
        # Get performance records
        performance_records = Performance.query.filter_by(
            customer_id=customer.id,
            group_id=group_id
        ).order_by(Performance.assessment_date.desc()).all()
        
        avg_score = sum([p.score for p in performance_records if p.score]) / len(performance_records) if performance_records else None
        
        member_performance.append({
            'member': member,
            'customer': customer,
            'attendance_rate': attendance_percentage,
            'total_sessions': total_sessions_possible,
            'sessions_attended': total_sessions_attended,
            'avg_score': avg_score,
            'recent_performances': performance_records[:3]
        })
    
    # Group history and events
    group_history = GroupHistory.query.filter_by(group_id=group_id).order_by(
        GroupHistory.created_at.desc()
    ).limit(20).all()
    
    # Get available customers for adding to group
    current_member_ids = [member[0].customer_id for member in members]
    available_customers = Customer.query.filter(
        ~Customer.id.in_(current_member_ids),
        Customer.status == 'active'
    ).all()
    
    # Progress timeline data for charts
    session_dates = [s.session_date for s in all_sessions if s.status == 'completed']
    session_dates.sort()
    
    # Monthly progress
    monthly_progress = {}
    for session_date in session_dates:
        month_key = session_date.strftime('%Y-%m')
        monthly_progress[month_key] = monthly_progress.get(month_key, 0) + 1
    
    # Overall group statistics
    group_stats = {
        'total_members': len(members),
        'total_sessions': total_sessions,
        'completed_sessions': completed_sessions,
        'cancelled_sessions': cancelled_sessions,
        'attendance_rate': round(attendance_rate, 1),
        'completion_rate': round((completed_sessions / total_sessions * 100) if total_sessions > 0 else 0, 1),
        'active_since_days': (datetime.now().date() - group.start_date).days,
        'avg_session_duration': sum([(datetime.combine(datetime.today(), s.end_time) - 
                                    datetime.combine(datetime.today(), s.start_time)).seconds / 60 
                                   for s in all_sessions]) / len(all_sessions) if all_sessions else 0
    }
    
    # Create detailed attendance matrix for table display
    attendance_matrix = {}
    completed_sessions = [s for s in all_sessions if s.status == 'completed']
    completed_sessions.sort(key=lambda x: x.session_date, reverse=True)
    
    # Limit to last 20 sessions for performance
    recent_completed_sessions = completed_sessions[:20]
    
    for session in recent_completed_sessions:
        session_key = f"{session.session_date}_{session.id}"
        attendance_matrix[session_key] = {
            'session': session,
            'attendance': {}
        }
        
        # Get attendance records for this session
        session_attendance = GroupAttendance.query.filter_by(group_session_id=session.id).all()
        for record in session_attendance:
            attendance_matrix[session_key]['attendance'][record.customer_id] = {
                'status': record.status,
                'notes': record.notes,
                'recorded_at': record.recorded_at
            }
    
    # Enhanced member performance with detailed attendance history
    enhanced_member_performance = []
    for member, customer in members:
        # Get all attendance records for this member
        member_attendance_records = db.session.query(GroupAttendance).join(GroupSession).filter(
            GroupSession.group_id == group_id,
            GroupAttendance.customer_id == customer.id,
            GroupSession.status == 'completed'
        ).order_by(GroupSession.session_date.desc()).all()
        
        # Calculate detailed statistics
        total_sessions = len(member_attendance_records)
        present_count = len([a for a in member_attendance_records if a.status == 'present'])
        late_count = len([a for a in member_attendance_records if a.status == 'late'])
        absent_count = len([a for a in member_attendance_records if a.status == 'absent'])
        excused_count = len([a for a in member_attendance_records if a.status == 'excused'])
        
        attendance_percentage = (present_count / total_sessions * 100) if total_sessions > 0 else 0
        
        # Get recent attendance (last 10 sessions)
        recent_attendance = member_attendance_records[:10]
        
        # Calculate attendance streak (consecutive present sessions)
        current_streak = 0
        for record in member_attendance_records:
            if record.status == 'present':
                current_streak += 1
            else:
                break
        
        # Get performance records
        performance_records = Performance.query.filter_by(
            customer_id=customer.id,
            group_id=group_id
        ).order_by(Performance.assessment_date.desc()).all()
        
        avg_score = sum([p.score for p in performance_records if p.score]) / len(performance_records) if performance_records else None
        
        enhanced_member_performance.append({
            'member': member,
            'customer': customer,
            'attendance_rate': attendance_percentage,
            'total_sessions': total_sessions,
            'present_count': present_count,
            'late_count': late_count,
            'absent_count': absent_count,
            'excused_count': excused_count,
            'current_streak': current_streak,
            'recent_attendance': recent_attendance,
            'avg_score': avg_score,
            'recent_performances': performance_records[:3],
            'attendance_records': member_attendance_records
        })
    
    return render_template('group_detail.html', 
                         group=group, 
                         members=members,
                         all_members=all_members,
                         upcoming_sessions=upcoming_sessions, 
                         recent_sessions=recent_sessions,
                         available_customers=available_customers,
                         member_performance=enhanced_member_performance,
                         group_history=group_history,
                         group_stats=group_stats,
                         monthly_progress=monthly_progress,
                         attendance_matrix=attendance_matrix,
                         recent_completed_sessions=recent_completed_sessions,
                         current_date=datetime.now().date())

@app.route('/groups/<int:group_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_group(group_id):
    group = Group.query.get_or_404(group_id)
    
    # Check permissions
    if current_user.role == 'instructor' and group.instructor_id != current_user.id:
        flash('Access denied.', 'error')
        return redirect(url_for('groups'))
    
    if request.method == 'POST':
        group.name = request.form['name']
        group.subject = request.form['subject']
        group.description = request.form.get('description', '')
        group.instructor_id = request.form['instructor_id']
        group.category_id = request.form.get('category_id') if request.form.get('category_id') else None
        group.max_students = int(request.form.get('max_students', 15))
        group.start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date()
        group.end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d').date() if request.form.get('end_date') else None
        group.notes = request.form.get('notes', '')
        group.status = request.form.get('status', 'active')
        
        # Update schedules - first delete existing ones
        GroupSchedule.query.filter_by(group_id=group.id).delete()
        
        # Add new schedules
        days = request.form.getlist('days[]')
        start_times = request.form.getlist('start_times[]')
        end_times = request.form.getlist('end_times[]')
        locations = request.form.getlist('locations[]')
        
        for i, day in enumerate(days):
            if day and i < len(start_times) and i < len(end_times):
                schedule = GroupSchedule(
                    group_id=group.id,
                    day_of_week=day,
                    start_time=datetime.strptime(start_times[i], '%H:%M').time(),
                    end_time=datetime.strptime(end_times[i], '%H:%M').time(),
                    location=locations[i] if i < len(locations) else ''
                )
                db.session.add(schedule)
        
        db.session.commit()
        flash('Group updated successfully!', 'success')
        return redirect(url_for('group_detail', group_id=group.id))
    
    instructors = User.query.filter_by(role='instructor', is_active=True).all()
    categories = CourseCategory.query.all()
    return render_template('edit_group.html', group=group, instructors=instructors, categories=categories)

@app.route('/groups/<int:group_id>/add_member', methods=['POST'])
@login_required
def add_group_member(group_id):
    group = Group.query.get_or_404(group_id)
    
    # Check permissions - instructors can only manage their own groups
    if current_user.role == 'instructor' and group.instructor_id != current_user.id:
        flash('Access denied. You can only manage your own groups.', 'error')
        return redirect(url_for('group_detail', group_id=group_id))
        
    group = Group.query.get_or_404(group_id)
    customer_id = request.form['customer_id']
    
    # Check if already a member
    existing_member = GroupMember.query.filter_by(
        group_id=group_id,
        customer_id=customer_id,
        status='active'
    ).first()
    
    if existing_member:
        flash('Customer is already a member of this group.', 'warning')
    else:
        # Check group capacity
        current_members = GroupMember.query.filter_by(group_id=group_id, status='active').count()
        if current_members >= group.max_students:
            flash('Group is at maximum capacity.', 'error')
        else:
            member = GroupMember(
                group_id=group_id,
                customer_id=customer_id,
                notes=request.form.get('notes', '')
            )
            db.session.add(member)
            db.session.commit()
            flash('Member added successfully!', 'success')
    
    return redirect(url_for('group_detail', group_id=group_id))

@app.route('/groups/<int:group_id>/members/<int:member_id>/remove', methods=['POST'])
@login_required
def remove_group_member(group_id, member_id):
    group = Group.query.get_or_404(group_id)
    
    # Check permissions - instructors can only manage their own groups
    if current_user.role == 'instructor' and group.instructor_id != current_user.id:
        flash('Access denied. You can only manage your own groups.', 'error')
        return redirect(url_for('group_detail', group_id=group_id))
        
    member = GroupMember.query.get_or_404(member_id)
    member.status = 'inactive'
    db.session.commit()
    flash('Member removed from group.', 'success')
    return redirect(url_for('group_detail', group_id=group_id))

@app.route('/groups/<int:group_id>/delete', methods=['POST'])
@login_required
def delete_group(group_id):
    # Only allow admin and customer_service roles to delete groups
    if current_user.role not in ['admin', 'customer_service']:
        flash('Access denied', 'error')
        return redirect(url_for('groups'))
    
    group = Group.query.get_or_404(group_id)
    
    # Check if group has any related data that would prevent deletion
    has_schedules = bool(group.schedules)
    has_members = bool(group.members)
    has_group_sessions = bool(group.group_sessions)
    has_performance_records = bool(group.performance_records)
    has_history_events = bool(group.history_events)
    
    # Check for attendance records through group sessions
    has_attendance_records = False
    if has_group_sessions:
        for session in group.group_sessions:
            if session.attendance_records:
                has_attendance_records = True
                break
    
    # If group has any related data, prevent deletion
    if (has_schedules or has_members or has_group_sessions or 
        has_performance_records or has_history_events or has_attendance_records):
        
        related_items = []
        if has_schedules:
            related_items.append(f"{len(group.schedules)} schedule(s)")
        if has_members:
            related_items.append(f"{len(group.members)} member(s)")
        if has_group_sessions:
            related_items.append(f"{len(group.group_sessions)} session(s)")
        if has_performance_records:
            related_items.append(f"{len(group.performance_records)} performance record(s)")
        if has_attendance_records:
            total_attendance = sum(len(session.attendance_records) for session in group.group_sessions)
            related_items.append(f"{total_attendance} attendance record(s)")
        if has_history_events:
            related_items.append(f"{len(group.history_events)} history event(s)")
        
        related_text = ", ".join(related_items)
        flash(f'Cannot delete group {group.name}. Group has {related_text}. Please remove or transfer these items first.', 'error')
        return redirect(url_for('groups'))
    
    try:
        # Log the deletion activity
        log_activity(
            entity_type='group',
            entity_id=group.id,
            action='deleted',
            user_id=current_user.id,
            description=f'Group {group.name} deleted',
            old_values=json.dumps({
                'name': group.name,
                'subject': group.subject,
                'description': group.description,
                'instructor_id': group.instructor_id,
                'max_students': group.max_students,
                'status': group.status,
                'start_date': group.start_date.isoformat() if group.start_date else None,
                'end_date': group.end_date.isoformat() if group.end_date else None,
                'created_at': group.created_at.isoformat() if group.created_at else None
            })
        )
        
        # Store group name for flash message before deletion
        group_name = group.name
        
        # Delete the group
        db.session.delete(group)
        db.session.commit()
        
        flash(f'Group {group_name} has been deleted successfully', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting group: {str(e)}', 'error')
    
    return redirect(url_for('groups'))

@app.route('/groups/<int:group_id>/sessions')
@login_required
def group_sessions(group_id):
    group = Group.query.get_or_404(group_id)
    
    # Check permissions
    if current_user.role == 'instructor' and group.instructor_id != current_user.id:
        flash('Access denied.', 'error')
        return redirect(url_for('groups'))
    
    sessions = GroupSession.query.filter_by(group_id=group_id).order_by(
        GroupSession.session_date.desc(), GroupSession.start_time.desc()
    ).all()
    
    return render_template('group_sessions.html', group=group, sessions=sessions)

@app.route('/groups/<int:group_id>/sessions/add', methods=['POST'])
@login_required
def add_group_session(group_id):
    group = Group.query.get_or_404(group_id)
    
    # Check permissions - instructors can only manage their own groups
    if current_user.role == 'instructor' and group.instructor_id != current_user.id:
        flash('Access denied. You can only schedule sessions for your own groups.', 'error')
        return redirect(url_for('group_sessions', group_id=group_id))
        
    group = Group.query.get_or_404(group_id)
    
    session = GroupSession(
        group_id=group_id,
        session_date=datetime.strptime(request.form['session_date'], '%Y-%m-%d').date(),
        start_time=datetime.strptime(request.form['start_time'], '%H:%M').time(),
        end_time=datetime.strptime(request.form['end_time'], '%H:%M').time(),
        topic=request.form.get('topic', ''),
        location=request.form.get('location', ''),
        notes=request.form.get('notes', '')
    )
    
    db.session.add(session)
    db.session.commit()
    flash('Session added successfully!', 'success')
    return redirect(url_for('group_sessions', group_id=group_id))

@app.route('/groups/sessions/<int:session_id>/attendance', methods=['GET', 'POST'])
@login_required
def group_session_attendance(session_id):
    session = GroupSession.query.get_or_404(session_id)
    
    if request.method == 'POST':
        # Clear existing attendance
        GroupAttendance.query.filter_by(group_session_id=session_id).delete()
        
        # Add new attendance records
        member_ids = request.form.getlist('member_ids[]')
        statuses = request.form.getlist('statuses[]')
        notes_list = request.form.getlist('notes[]')
        
        for i, member_id in enumerate(member_ids):
            if i < len(statuses):
                attendance = GroupAttendance(
                    group_session_id=session_id,
                    customer_id=member_id,
                    status=statuses[i],
                    notes=notes_list[i] if i < len(notes_list) else ''
                )
                db.session.add(attendance)
        
        session.status = 'completed'
        db.session.commit()
        flash('Attendance recorded successfully!', 'success')
        return redirect(url_for('group_sessions', group_id=session.group_id))
    
    # Get group members
    members = db.session.query(GroupMember, Customer).join(Customer).filter(
        GroupMember.group_id == session.group_id,
        GroupMember.status == 'active'
    ).all()
    
    # Get existing attendance
    existing_attendance = {}
    attendance_records = GroupAttendance.query.filter_by(group_session_id=session_id).all()
    for record in attendance_records:
        existing_attendance[record.customer_id] = record
    
    return render_template('group_attendance.html', session=session, members=members, 
                         existing_attendance=existing_attendance)

@app.route('/performance/add', methods=['POST'])
@login_required
def add_performance_record():
    """Add a performance/assessment record for a student"""
    # منع المدرسين من إضافة تقييمات
    if current_user.role == 'instructor':
        flash('Access denied. Instructors cannot add assessments.', 'error')
        return redirect(url_for('group_detail', group_id=request.form.get('group_id', 1)))
        
    try:
        performance = Performance(
            customer_id=request.form['customer_id'],
            group_id=request.form['group_id'],
            assessment_date=datetime.strptime(request.form['assessment_date'], '%Y-%m-%d').date(),
            score=float(request.form['score']) if request.form.get('score') else None,
            grade=request.form.get('grade', '').upper() if request.form.get('grade') else None,
            assessment_type=request.form['assessment_type'],
            notes=request.form.get('notes', ''),
            created_by_id=current_user.id
        )
        
        db.session.add(performance)
        
        # Log this activity
        customer = Customer.query.get(request.form['customer_id'])
        group = Group.query.get(request.form['group_id'])
        log_customer_event(
            customer_id=customer.id,
            event_type='assessment_added',
            description=f'Assessment added: {performance.assessment_type} in {group.name}',
            event_data={
                'group_id': group.id,
                'score': performance.score,
                'grade': performance.grade,
                'assessment_type': performance.assessment_type
            }
        )
        
        log_group_event(
            group_id=group.id,
            event_type='assessment_added',
            description=f'Assessment added for {customer.first_name} {customer.last_name}',
            event_data={
                'customer_id': customer.id,
                'score': performance.score,
                'assessment_type': performance.assessment_type
            }
        )
        
        db.session.commit()
        flash('Assessment record added successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding assessment: {str(e)}', 'error')
    
    return redirect(url_for('group_detail', group_id=request.form['group_id']))

@app.route('/communication/add', methods=['POST'])
@login_required
def add_communication_record():
    """Add a communication record for a customer"""
    try:
        communication = Communication(
            customer_id=request.form['customer_id'],
            communication_type=request.form['communication_type'],
            direction=request.form['direction'],
            subject=request.form.get('subject', ''),
            content=request.form['content'],
            status=request.form.get('status', 'completed'),
            response_received=bool(request.form.get('response_received')),
            created_by_id=current_user.id
        )
        
        db.session.add(communication)
        
        # Log this activity
        customer = Customer.query.get(request.form['customer_id'])
        log_customer_event(
            customer_id=customer.id,
            event_type='communication_logged',
            description=f'{communication.communication_type.title()} communication: {communication.subject or communication.content[:50]}',
            event_data={
                'type': communication.communication_type,
                'direction': communication.direction,
                'status': communication.status
            }
        )
        
        db.session.commit()
        flash('Communication record added successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding communication: {str(e)}', 'error')
    
    return redirect(url_for('customer_detail', customer_id=request.form['customer_id']))

def create_app():
    """Application factory pattern for easier testing and deployment"""
    return app

def init_database():
    """Initialize database with default data"""
    with app.app_context():
        # Create tables
        db.create_all()
        
        # Create default admin user if it doesn't exist
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin = User(
                username='admin',
                email='admin@example.com',
                password_hash=generate_password_hash('admin123'),
                role='admin',
                first_name='System',
                last_name='Administrator'
            )
            db.session.add(admin)
            db.session.commit()
            print("Default admin user created: username='admin', password='admin123'")
        
        # Create default categories if they don't exist
        default_categories = [
            {'name': 'Competitions', 'description': 'Programming competitions and contests'},
            {'name': 'Camps', 'description': 'Intensive training camps and bootcamps'},
            {'name': 'Courses', 'description': 'Regular educational courses and classes'}
        ]
        
        for cat_data in default_categories:
            existing_category = CourseCategory.query.filter_by(name=cat_data['name']).first()
            if not existing_category:
                category = CourseCategory(
                    name=cat_data['name'],
                    description=cat_data['description']
                )
                db.session.add(category)
                print(f"Created category: {cat_data['name']}")
        
        db.session.commit()

if __name__ == '__main__':
    # Only run if called directly (development mode)
    init_database()
    
    # Get configuration from environment
    host = os.environ.get('HOST', '127.0.0.1')
    port = int(os.environ.get('PORT', 8002))
    debug = os.environ.get('FLASK_ENV', 'development') == 'development'
    
    app.run(host=host, port=port, debug=debug) 